"""
Dark Filaments — Simulator v1.3

v1.3 adds Tier 2. T1 sheets remain unchanged from v1.2.1.

Architecture note (Option B): T2 is a separate sheet (UpgradeSimT2) parallel
to UpgradeSim. It takes T1's exit state (mass, MPS-floor, MPC-floor, AC/s-floor,
active multipliers from owned T1 one-shots) as carryover inputs at the top of
the sheet. T1 upgrades cannot be releveled in T2; their contributions become
a "carry-over income floor" that T2 builds on top of.

Carryover scenarios:
  - Completion-exit (default): all T1 completionists owned (Mag=5, FP=1).
    Active multipliers: ×1.25 M/sec (OR), ×1.20 M/click (FP).
    Mass surplus typical ~315 at 100 cpm.
  - Threshold-exit: cohesion 1.0 met but FP/Mag not maxed.
    Active multipliers: ×1.25 M/sec (OR), no FP click bonus.
    Mass surplus typical ~80 at 100 cpm.
  Both are configurable via Parameters; default is Completion-exit.

T2 upgrades (9 total, v3 spec):
  Stackables:    Stellar Kinematics, Local Bubble, Microlensing,
                 Roche Lobe Overflow (first APS source), Brown Dwarf (max-5 completionist)
  One-shots:     Binary Partner, Peculiar Velocity, Open Cluster (T3 gate),
                 Moving Group (completionist; +6 M/s base AND ×1.20 M/click)
  Synergies:     A. Binary Partner → Microlensing × 1.5  (one-shot, flat)
                 B. Roche Lobe Overflow → Local Bubble × 1.05^L  (stackable per-level)
                 C. Brown Dwarf → Roche Lobe Overflow × 1.10^L  (stackable cross-stat)

VPC for synergy *providers* now includes the value they add to their target,
not just their own self-contribution. This fixes a known formula gap that
would otherwise underweight RLO and BD.

Strategy carries over from T1: greedy VPC + save_vpc_threshold = 1.5
+ post-cohesion focus.

T1 strategy carried unchanged. The save_ratio parameter is left in place
for backward compatibility but is not used by the action formula.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.comments import Comment

# ---------- styling ----------
FONT_NAME = "Arial"

H1 = Font(name=FONT_NAME, size=16, bold=True, color="111827")
H2 = Font(name=FONT_NAME, size=12, bold=True, color="1F2937")
H3 = Font(name=FONT_NAME, size=11, bold=True, color="374151")
BODY = Font(name=FONT_NAME, size=10, color="1F2937")
INPUT_FONT = Font(name=FONT_NAME, size=10, color="1E40AF", bold=True)
FORMULA_FONT = Font(name=FONT_NAME, size=10, color="000000")
LINK_FONT = Font(name=FONT_NAME, size=10, color="047857")
NOTE_FONT = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
EVENT_FONT = Font(name=FONT_NAME, size=9, italic=True, color="7C2D12")

INPUT_FILL = PatternFill("solid", start_color="FEF3C7")
HEADER_FILL = PatternFill("solid", start_color="1F2937")
SECTION_FILL = PatternFill("solid", start_color="F3F4F6")
DERIVED_FILL = PatternFill("solid", start_color="ECFDF5")  # light green for derived
CARRY_FILL = PatternFill("solid", start_color="E0E7FF")   # light blue for T1 carryover

THIN = Side(border_style="thin", color="D1D5DB")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Alignments — separate "wrap" and "no-wrap" variants
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_NOWRAP = Alignment(horizontal="left", vertical="center", wrap_text=False)
RIGHT = Alignment(horizontal="right", vertical="center")


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def label(ws, row, col, text, comment=None, alignment=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BODY
    c.alignment = alignment or LEFT_NOWRAP
    if comment:
        c.comment = Comment(comment, "Claude")


def input_cell(ws, row, col, value, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = INPUT_FONT
    c.fill = INPUT_FILL
    c.alignment = RIGHT
    c.border = BORDER_ALL
    if number_format:
        c.number_format = number_format
    return c


def formula_cell(ws, row, col, formula, number_format=None, font=None,
                 fill=None, alignment=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.font = font or FORMULA_FONT
    c.alignment = alignment or RIGHT
    c.border = BORDER_ALL
    if number_format:
        c.number_format = number_format
    if fill:
        c.fill = fill
    return c


def header_cell(ws, row, col, text, comment=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER_ALL
    if comment:
        c.comment = Comment(comment, "Claude")
    return c


# ---------- workbook ----------
wb = Workbook()
wb.remove(wb.active)


# ============================================================
# README
# ============================================================
readme = wb.create_sheet("README")
set_widths(readme, {"A": 110})

# Title — explicitly NO wrapping, single line
title = readme.cell(row=1, column=1, value="Dark Filaments — Simulator v1.3")
title.font = H1
title.alignment = LEFT_NOWRAP

subtitle = readme.cell(row=2, column=1,
                       value="Adds Tier 2 in a separate sheet (UpgradeSimT2) that consumes T1's exit state as carryover. T1 sheet is unchanged from v1.2.1. Greedy VPC + save threshold + post-cohesion focus strategy carries over.")
subtitle.font = NOTE_FONT
subtitle.alignment = LEFT_NOWRAP

content = [
    ("", None),
    ("The three core stats", "h2"),
    ("• MPC — Mass per Click — affects active click income AND autoclicker output", "body"),
    ("• MPS — Mass per Second — passive income rate", "body"),
    ("• APS — Autoclicks per Second — fake clicks per second from autoclickers", "body"),
    ("Game tick = 1 second. Total income/sec = (cpm/60 × MPC × active) + (APS × MPC) + MPS.", "body"),
    ("Simulator measures in 10-second chunks (10 game ticks per sim row).", "body"),
    ("", None),
    ("Unified upgrade schema", "h2"),
    ("Every upgrade has the same fields. Most fields will be 0 (additive) or 1.0 (multiplicative) "
     "for any given upgrade — only the ones that actually do something get filled in.", "body"),
    ("", None),
    ("Per-stat fields (one set each for M/sec, M/click, AC/sec):", "h3"),
    ("• Base — constant value when level ≥ 1, before any multipliers", "body"),
    ("• + per lvl — linear growth: each level adds this amount", "body"),
    ("• × self per lvl — exponential growth: each level multiplies THIS upgrade by this factor", "body"),
    ("", None),
    ("Global multiplier fields (apply to ALL upgrades' contributions of that type):", "h3"),
    ("• × all M/sec per lvl — buffs all passive sources", "body"),
    ("• × all M/click per lvl — buffs all click value sources", "body"),
    ("• × all AC speed per lvl — buffs all autoclicker rates", "body"),
    ("", None),
    ("Contribution formula:", "h3"),
    ("contribution(level N, upgrade U) = IF(N = 0, 0, (U.base + N × U.additive) × U.self_mult^N)", "body"),
    ("Total stat = SUM(self contributions) × PRODUCT(global multipliers ^ levels)", "body"),
    ("", None),
    ("How current T1 maps to the schema", "h2"),
    ("• Solar Wind — additive M/sec only: + per lvl = 0.02. Pure linear passive.", "body"),
    ("• Asteroid Belt — additive M/sec only: + per lvl = 0.07. Pure linear passive.", "body"),
    ("• Stellar Coupling — additive M/click only: + per lvl = 0.32. Pure linear click.", "body"),
    ("• Orbital Resonance — × all M/sec per lvl = 1.25, max levels = 1. One-shot global passive multiplier.", "body"),
    ("• Heliopause — all effects 0/1.0, max levels = 1, cohesion 0.6. Pure cohesion contribution.", "body"),
    ("", None),
    ("Try a non-linear variant", "h2"),
    ("Want to test 'each Solar Wind level grows itself by 5%'? Set:", "body"),
    ("  Solar Wind: base = 0.02, + per lvl = 0, × self M/sec per lvl = 1.05", "body"),
    ("That makes level 5 produce (0.02 + 0) × 1.05^5 = 0.0255 M/sec — 27% more than linear at the same level.", "body"),
    ("Or for a pure black-hole-style upgrade: base = 1.0, + per lvl = 0, × self per lvl = 1.10. "
     "Level 10 = 1.0 × 1.10^10 = 2.59 — slow start, runaway end.", "body"),
    ("", None),
    ("Tier transitions", "h2"),
    ("Cohesion-gated only. Available when cohesion ≥ tier_transition_cost. "
     "Player buys cohesion-rewarding one-shots in cheapest-first order until they have enough. "
     "No 'gates' field needed — the cohesion math takes care of ordering.", "body"),
    ("", None),
    ("Carryover model (locked, unchanged from v0.7+)", "h2"),
    ("Effects from earlier tiers persist forever. You cannot buy MORE of a previous tier's "
     "upgrades after transitioning. Each tier's purchase loop is self-contained.", "body"),
    ("", None),
    ("Tab guide", "h2"),
    ("Parameters    — Master knobs. Includes T2 carryover block (mode toggle + per-mode carry values).", "body"),
    ("Upgrades      — Per-upgrade definitions in the unified schema. T1 rows 6-12, T2 rows 13-21.", "body"),
    ("TimeBudget    — Per-tier minute targets and click share targets.", "body"),
    ("UpgradeSim    — Tier 1 micro simulation (unchanged from v1.2.1).", "body"),
    ("Curves        — T1 purchase cadence and click share diagnostics.", "body"),
    ("UpgradeSimT2  — Tier 2 micro simulation. Carryover block at top reflects T1 exit state.", "body"),
    ("CurvesT2      — T2 purchase cadence and active/auto/passive income split.", "body"),
    ("", None),
    ("What got dropped from earlier versions", "h2"),
    ("• income_ratio_total parameter (we never landed on a value that meant anything)", "body"),
    ("• Macro income target curves on TierPlan (replaced by simpler TimeBudget)", "body"),
    ("• Macro 1-min Simulation tab (was diagnostic for value targets we no longer track)", "body"),
    ("• Events/min density parameters (was never properly defined as a metric)", "body"),
    ("• Rigid upgrade type system (replaced by unified schema)", "body"),
    ("• 'Gates' field on upgrades (cohesion-gating handles transitions instead)", "body"),
    ("", None),
    ("Color conventions", "h2"),
    ("Yellow = editable input. Light green = derived live calculation. "
     "Black text = formula. Green text = cross-tab reference.", "body"),
]

row = 3
for text, kind in content:
    c = readme.cell(row=row, column=1, value=text)
    if kind == "h2":
        c.font = H2
        c.fill = SECTION_FILL
        c.alignment = LEFT_NOWRAP
        readme.row_dimensions[row].height = 22
    elif kind == "h3":
        c.font = H3
        c.alignment = LEFT_NOWRAP
        readme.row_dimensions[row].height = 18
    elif kind == "body":
        c.font = BODY
        c.alignment = LEFT_WRAP
        readme.row_dimensions[row].height = max(18, 14 * (1 + len(text) // 95))
    row += 1


# ============================================================
# Parameters
# ============================================================
params = wb.create_sheet("Parameters")
set_widths(params, {"A": 50, "B": 18, "C": 60})

ttitle = params.cell(row=1, column=1, value="Parameters")
ttitle.font = H1
ttitle.alignment = LEFT_NOWRAP
sub = params.cell(row=2, column=1,
                  value="Edit yellow cells. Everything else updates automatically.")
sub.font = NOTE_FONT
sub.alignment = LEFT_NOWRAP

p_row = 4
PARAM_ROWS = {}


def add_param(name, label_text, value, note, fmt=None):
    global p_row
    label(params, p_row, 1, label_text, alignment=LEFT_NOWRAP)
    input_cell(params, p_row, 2, value, number_format=fmt)
    nc = params.cell(row=p_row, column=3, value=note)
    nc.font = NOTE_FONT
    nc.alignment = LEFT_WRAP
    PARAM_ROWS[name] = p_row
    p_row += 1


def add_section(text):
    global p_row
    sc = params.cell(row=p_row, column=1, value=text)
    sc.font = H2
    sc.fill = SECTION_FILL
    sc.alignment = LEFT_NOWRAP
    params.merge_cells(start_row=p_row, start_column=1, end_row=p_row, end_column=3)
    p_row += 1


add_section("Time budget")
add_param("total_minutes", "Total target playtime (minutes)",
          595, "Casual pace at 45 cpm. 60 cpm players will finish closer to ~480 min.", "0")
add_param("num_tiers", "Number of tiers", 10, "Solar System through Causal Horizon.", "0")
add_param("tier_time_growth", "Tier real-time growth rate", 1.30,
          "Each tier takes this many times as long as the previous one.", "0.00")

add_section("Click & income")
add_param("base_click_value", "Base click value (Mass per click, no upgrades)",
          1.0, "The unit of the economy. Each click yields this much Mass with zero upgrades.", "0.00")
add_param("active_clicks_per_min", "Average clicks per minute when active",
          100, "Reflects autoclicker / engaged play. Prototype dynamically reports actual cpm.", "0")
add_param("active_play_fraction", "Active play fraction (0–1)",
          1.00, "Prototype reports cpm in real time, so this stays at 1.0; AFK time is captured by cpm dropping.", "0.00")

add_section("Click importance curve (U-shape, diagnostic only)")
add_param("click_share_T1", "Click share at Tier 1", 0.60, "Clicks dominate at start.", "0%")
add_param("click_share_T5_min", "Click share minimum (Tier 5 dip)", 0.05,
          "Crescendo — chain reactions handle nearly everything.", "0%")
add_param("click_share_T9", "Click share late game (Tier 9)", 0.65, "'Hanging on.'", "0%")
add_param("click_share_T10", "Click share endgame (Tier 10)", 0.70, "Click becomes ritual.", "0%")

add_section("Cohesion (gates tier transitions)")
add_param("cohesion_T1_to_T2", "Cohesion cost: T1 → T2 transition", 1.0, "Tutorial-friendly.", "0.0")
add_param("cohesion_growth", "Cohesion cost growth per tier", 2.50, "Geometric growth.", "0.00")

add_section("Simulator strategy")
add_param("save_vpc_threshold", "VPC save threshold",
          1.5, "Save mode triggers when next-target VPC > (best-affordable VPC × this). Lower = saves harder.", "0.0")

add_section("Tier 2 carryover (from T1 exit state)")
add_param("t2_carry_mode", "T2 carryover mode (1=Completion, 0=Threshold)",
          1, "1: T1 fully completed (Mag=5, FP=1). 0: T1 cohesion met but completionists not maxed.", "0")
add_param("t2_carry_mass_completion", "Carry mass at T1 Completion-exit",
          315, "Mass surplus when T1 finishes the completion path at 100 cpm.", "0")
add_param("t2_carry_mass_threshold", "Carry mass at T1 Threshold-exit",
          80, "Mass surplus when T1 just barely meets cohesion (Mag~2, FP=0).", "0")
add_param("t2_carry_mps_completion", "Carry MPS floor at Completion-exit",
          10.75, "T1 final MPS contribution. Persists into T2 as a floor.", "0.00")
add_param("t2_carry_mps_threshold", "Carry MPS floor at Threshold-exit",
          5.75, "T1 partial MPS at threshold. Persists into T2 as a floor.", "0.00")
add_param("t2_carry_mpc_completion", "Carry MPC floor at Completion-exit",
          4.98, "T1 final MPC contribution including base click value × all multipliers.", "0.00")
add_param("t2_carry_mpc_threshold", "Carry MPC floor at Threshold-exit",
          4.15, "T1 partial MPC including base click and active multipliers (no FP).", "0.00")
add_param("t2_carry_aps", "Carry APS floor", 0.0,
          "T1 has no APS sources. Stays 0 until that changes.", "0.00")
add_param("t2_carry_all_mps_completion", "Carry × all M/sec at Completion-exit",
          1.25, "OR active (×1.25). Continues to multiply T2 MPS contributions.", "0.00")
add_param("t2_carry_all_mps_threshold", "Carry × all M/sec at Threshold-exit",
          1.25, "OR active at threshold-exit (it's bought to reach cohesion 1.0).", "0.00")
add_param("t2_carry_all_mpc_completion", "Carry × all M/click at Completion-exit",
          1.20, "FP active (×1.20). Continues to multiply T2 MPC contributions.", "0.00")
add_param("t2_carry_all_mpc_threshold", "Carry × all M/click at Threshold-exit",
          1.00, "FP not yet owned at threshold-exit.", "0.00")
add_param("t2_carry_all_aps", "Carry × all AC speed", 1.00,
          "No T1 sources affect AC speed.", "0.00")

add_section("Offline progression")
add_param("offline_rate", "Offline progression rate (1.0 = full)",
          1.0, "Full offline accumulation; queued events fire on return.", "0.0%")

add_section("Tier names")
TIER_NAMES_DEFAULT = [
    "Solar System", "Stellar Neighborhood", "Galactic Arm", "Galaxy",
    "Local Group", "Galactic Cluster", "Supercluster", "Filament",
    "Cosmic Web", "Causal Horizon",
]
TIER_NAME_START_ROW = p_row
for i, name in enumerate(TIER_NAMES_DEFAULT, start=1):
    label(params, p_row, 1, f"Tier {i} name")
    input_cell(params, p_row, 2, name)
    note = ""
    if i == 1: note = "Tutorial."
    elif i == 2: note = "First sense of scale. Introduces autoclickers."
    elif i == 5: note = "PEAK. Crescendo."
    elif i == 10: note = "Endgame."
    if note:
        nc = params.cell(row=p_row, column=3, value=note)
        nc.font = NOTE_FONT
        nc.alignment = LEFT_WRAP
    p_row += 1


def P(name):
    return f"Parameters!$B${PARAM_ROWS[name]}"


# ============================================================
# Upgrades — unified schema
# ============================================================
upg = wb.create_sheet("Upgrades")
upg_widths = {
    "A": 6,   # Tier
    "B": 22,  # Name
    "C": 11,  # Init cost
    "D": 11,  # Cost growth
    "E": 11,  # Max levels
    "F": 13,  # Cohesion reward
    "G": 11,  # Base M/sec
    "H": 11,  # + M/sec/lvl
    "I": 13,  # × self M/sec/lvl
    "J": 11,  # Base M/click
    "K": 11,  # + M/click/lvl
    "L": 13,  # × self M/click/lvl
    "M": 11,  # Base AC/sec
    "N": 11,  # + AC/sec/lvl
    "O": 13,  # × self AC/sec/lvl
    "P": 13,  # × all M/sec/lvl
    "Q": 13,  # × all M/click/lvl
    "R": 13,  # × all AC speed/lvl
    "S": 13,  # Completionist?
    "T": 18,  # Synergy target
    "U": 13,  # Synergy mult/lvl
    "V": 50,  # Description
}
set_widths(upg, upg_widths)

ut = upg.cell(row=1, column=1, value="Upgrades — unified schema")
ut.font = H1
ut.alignment = LEFT_NOWRAP

us = upg.cell(row=2, column=1,
              value="Each row is one upgrade. Yellow cells editable. "
                    "Most cells will be 0 (additive) or 1.0 (multiplicative) — only the levers that "
                    "actually do something get filled in. Completionist = missable upgrade. Synergy "
                    "target/mult lets one upgrade buff another's self-contribution geometrically with level.")
us.font = NOTE_FONT
us.alignment = LEFT_NOWRAP
upg.row_dimensions[2].height = 18

# Group bar in row 4
group_specs = [
    (1, 2, "Identity"),
    (3, 6, "Cost & limits"),
    (7, 9, "Self M/sec"),
    (10, 12, "Self M/click"),
    (13, 15, "Self AC/sec"),
    (16, 18, "Global multipliers"),
    (19, 19, "Completion"),
    (20, 21, "Synergy"),
    (22, 22, "Description"),
]
for col_start, col_end, gtext in group_specs:
    gc = upg.cell(row=4, column=col_start, value=gtext)
    gc.font = H3
    gc.fill = SECTION_FILL
    gc.alignment = CENTER
    gc.border = BORDER_ALL
    if col_end > col_start:
        upg.merge_cells(start_row=4, start_column=col_start,
                        end_row=4, end_column=col_end)

# Row 5 = field headers
HEADERS = [
    ("Tier", None),
    ("Name", None),
    ("Init cost", "Mass cost of the FIRST purchase. Subsequent purchases grow by Cost growth ^ level."),
    ("Cost growth", "Cost multiplier per level. 1.0 = flat (one-shots use this). 1.15+ = standard idle compounding."),
    ("Max levels", "Cap on purchases. 1 = one-shot. 99 = effectively unlimited."),
    ("Cohesion +", "Cohesion granted per purchase. Sums across all purchases — used to gate tier transitions."),
    ("Base", "Base M/sec contribution when level ≥ 1, before any growth or multipliers."),
    ("+ /lvl", "Additive M/sec per level. Pure linear growth lever."),
    ("× self /lvl", "Self-multiplicative growth on THIS upgrade's M/sec only."),
    ("Base", "Base M/click contribution when level ≥ 1."),
    ("+ /lvl", "Additive M/click per level."),
    ("× self /lvl", "Self-multiplicative growth on THIS upgrade's M/click only."),
    ("Base", "Base AC/sec contribution when level ≥ 1."),
    ("+ /lvl", "Additive autoclicks per second per level."),
    ("× self /lvl", "Self-multiplicative growth on THIS upgrade's AC/sec."),
    ("× all M/sec /lvl", "Per-level multiplier on ALL passive sources across all upgrades."),
    ("× all M/click /lvl", "Per-level multiplier on ALL click value sources."),
    ("× all AC speed /lvl", "Per-level multiplier on ALL autoclicker rates."),
    ("Completionist?", "1 = missable (max < 99 AND cohesion = 0). Sim under 'completion path' won't transition until these are owned."),
    ("Synergy target", "Name of another upgrade whose self-contribution is multiplied by this upgrade. Empty = no synergy."),
    ("Synergy mult/lvl", "Multiplier applied to the targeted upgrade. Raised to this upgrade's level (compounds with multiple levels of provider)."),
    ("Description voice", None),
]
for col, (h, c) in enumerate(HEADERS, start=1):
    cell = header_cell(upg, 5, col, h, comment=c)


# Tier 1 upgrades. Tuple format:
# (tier, name, init_cost, cost_growth, max_lvl, coh_reward,
#  base_mps, add_mps, self_mps,
#  base_mpc, add_mpc, self_mpc,
#  base_aps, add_aps, self_aps,
#  all_mps, all_mpc, all_aps,
#  completionist, description)
# completionist = 1 means a non-cohesion limited/one-shot — missable.
# Tier 1 upgrades. Tuple format:
# (tier, name, init_cost, cost_growth, max_lvl, coh_reward,
#  base_mps, add_mps, self_mps,
#  base_mpc, add_mpc, self_mpc,
#  base_aps, add_aps, self_aps,
#  all_mps, all_mpc, all_aps,
#  completionist, synergy_target, synergy_mult, description)
# completionist = 1 means a non-cohesion limited/one-shot — missable.
# synergy_target = name of the upgrade this one buffs (or "" for none).
# synergy_mult = multiplier raised to this upgrade's level for the targeted upgrade's self-contribution.
TIER_1_UPGRADES = [
    (1, "Solar Wind",
     7, 1.15, 99, 0.0,
     0, 0.080, 1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "Charged particles drift outward and return with company. We are pulling more than we used to."),
    (1, "Asteroid Belt",
     20, 1.15, 99, 0.0,
     0, 0.200, 1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "The belt yields. Iron, ice, the slow gravel of the early system. Each rock finds us."),
    (1, "Stellar Coupling",
     22, 1.40, 99, 0.0,
     0, 0,     1.0,
     0, 0.350, 1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "We pull harder. The center holds tighter."),
    (1, "Magnetosphere",
     80, 2.00, 5, 0.0,
     0, 1.000, 1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     1, "", 1.0,
     "The system's invisible shell. Charged particles arc and return. We catch what would have escaped."),
    (1, "Orbital Resonance",
     240, 1.0, 1, 0.4,
     0, 0,     1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.25, 1.0, 1.0,
     0, "", 1.0,
     "Periods align. The system breathes in time with us. Everything we touch becomes synchronous."),
    (1, "Heliopause",
     575, 1.0, 1, 0.6,
     0, 0,     1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "Stellar Coupling", 1.5,
     "We have reached the edge of our influence. Beyond it, the rest of the galaxy waits."),
    (1, "First Photons",
     850, 1.0, 1, 0.0,
     1.0, 0,   1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.20, 1.0,
     1, "", 1.0,
     "Light, finally. The first photons leave the surface and find us. Everything quickens."),
]

# Tier 2 upgrades — same tuple format as T1
# v3-locked numbers. Synergy column already supports per-level multipliers
# via POWER(provider_mult, provider_level), so:
#   Binary Partner (one-shot, level 1) → Microlensing × 1.5^1 = ×1.5 flat
#   Roche Lobe Overflow (stackable) → Local Bubble × 1.05^L_RLO  (kind B)
#   Brown Dwarf (stackable) → Roche Lobe Overflow × 1.10^L_BD   (kind C)
TIER_2_UPGRADES = [
    # 1. Stellar Kinematics — passive stackable, additive M/sec only
    (2, "Stellar Kinematics",
     60, 1.15, 99, 0.0,
     0, 0.65,  1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "Stars drift in measured patterns. We learn to read the slow vectors of the neighborhood."),
    # 2. Local Bubble — passive stackable, larger additive M/sec, target of synergy B
    (2, "Local Bubble",
     180, 1.15, 99, 0.0,
     0, 1.80,  1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "An emptier pocket of the galaxy. Our reach extends through its low-density walls."),
    # 3. Microlensing — click stackable, target of synergy A
    (2, "Microlensing",
     200, 1.40, 99, 0.0,
     0, 0,     1.0,
     0, 2.80,  1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "Distant light bends around mass we cannot see. Each event is a brief annunciation."),
    # 4. Roche Lobe Overflow — first APS source, target of synergy C, provider of synergy B
    (2, "Roche Lobe Overflow",
     350, 1.50, 99, 0.0,
     0, 0,     1.0,
     0, 0,     1.0,
     0, 0.10,  1.0,
     1.0, 1.0, 1.0,
     0, "Local Bubble", 1.05,
     "A companion star's outer envelope spills across the gravitational saddle. Material flows to us automatically."),
    # 5. Brown Dwarf — completionist stackable max-5, provider of synergy C
    (2, "Brown Dwarf",
     2000, 2.00, 5, 0.0,
     0, 4.00,  1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     1, "Roche Lobe Overflow", 1.10,
     "A failed star, never quite hot enough. We catalog it. Its weak gravity helps the lobes spill faster."),
    # 6. Binary Partner — one-shot, provider of synergy A
    (2, "Binary Partner",
     1400, 1.0, 1, 0.6,
     0, 0,     1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "Microlensing", 1.5,
     "We pair. The two-body orbit is mathematics made visible. Lensing events double when seen from a moving baseline."),
    # 7. Peculiar Velocity — one-shot, ×1.30 to all M/sec
    (2, "Peculiar Velocity",
     2200, 1.0, 1, 0.9,
     0, 0,     1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.30, 1.0, 1.0,
     0, "", 1.0,
     "Above the Hubble flow we have our own motion. Everything drifts a little faster toward us."),
    # 8. Open Cluster — one-shot, T3 transition gate (cohesion 1.0)
    (2, "Open Cluster",
     3400, 1.0, 1, 1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.0, 1.0,
     0, "", 1.0,
     "Loose stars share a birth and a direction. We are no longer alone in our trajectory."),
    # 9. Moving Group — completionist one-shot, +6 base M/sec AND ×1.20 M/click
    # Note on synergies: this upgrade ADDS to base M/sec via a flat 6.0 in the base column,
    # which means level=0 yields 0 contribution, level=1 yields 6.0. That's how T1 First Photons
    # encoded its base 1.0 M/sec. Same pattern.
    (2, "Moving Group",
     14000, 1.0, 1, 0.0,
     6.0, 0,   1.0,
     0, 0,     1.0,
     0, 0,     1.0,
     1.0, 1.20, 1.0,
     1, "", 1.0,
     "The whole pocket drifts together — Hyades, Pleiades, kin we never named. We feel the current."),
]

UPG_DATA_START_ROW = 6
UPG_ROW = {}  # name -> row number on Upgrades tab
ALL_TIER_UPGRADES = TIER_1_UPGRADES + TIER_2_UPGRADES
for i, u in enumerate(ALL_TIER_UPGRADES):
    r = UPG_DATA_START_ROW + i
    UPG_ROW[u[1]] = r
    # Identity
    formula_cell(upg, r, 1, u[0], number_format="0", font=INPUT_FONT, fill=INPUT_FILL)
    nc = upg.cell(row=r, column=2, value=u[1])
    nc.font = INPUT_FONT
    nc.fill = INPUT_FILL
    nc.alignment = LEFT_NOWRAP
    nc.border = BORDER_ALL
    # Cost & limits
    formula_cell(upg, r, 3, u[2], number_format="0", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 4, u[3], number_format="0.00", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 5, u[4], number_format="0", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 6, u[5], number_format="0.00", font=INPUT_FONT, fill=INPUT_FILL)
    # Self M/sec
    formula_cell(upg, r, 7, u[6], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 8, u[7], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 9, u[8], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    # Self M/click
    formula_cell(upg, r, 10, u[9], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 11, u[10], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 12, u[11], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    # Self AC/sec
    formula_cell(upg, r, 13, u[12], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 14, u[13], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 15, u[14], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    # Global multipliers
    formula_cell(upg, r, 16, u[15], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 17, u[16], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    formula_cell(upg, r, 18, u[17], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    # Completionist flag (col 19)
    formula_cell(upg, r, 19, u[18], number_format="0", font=INPUT_FONT, fill=INPUT_FILL)
    # Synergy target (col 20) — string
    sc = upg.cell(row=r, column=20, value=u[19])
    sc.font = INPUT_FONT
    sc.fill = INPUT_FILL
    sc.alignment = LEFT_NOWRAP
    sc.border = BORDER_ALL
    # Synergy mult/lvl (col 21)
    formula_cell(upg, r, 21, u[20], number_format="0.000", font=INPUT_FONT, fill=INPUT_FILL)
    # Description (col 22)
    dc = upg.cell(row=r, column=22, value=u[21])
    dc.font = NOTE_FONT
    dc.alignment = LEFT_WRAP

upg.freeze_panes = "C6"


# ============================================================
# TimeBudget — replaces TierPlan
# ============================================================
plan = wb.create_sheet("TimeBudget")
plan_widths = {"A": 6, "B": 22, "C": 12, "D": 14, "E": 12, "F": 14, "G": 36}
set_widths(plan, plan_widths)

pt = plan.cell(row=1, column=1, value="Time budget per tier")
pt.font = H1
pt.alignment = LEFT_NOWRAP

ps = plan.cell(row=2, column=1,
               value="Time targets and click share targets per tier. Derived from Parameters.")
ps.font = NOTE_FONT
ps.alignment = LEFT_NOWRAP

PLAN_HEADERS = [
    ("Tier", None),
    ("Name", None),
    ("Min/tier", "Real-time minutes target for this tier."),
    ("Cum min", "Cumulative minutes through end of this tier."),
    ("Click %", "Diagnostic target — what fraction of income should come from active clicking."),
    ("Cohesion cost", "Cohesion needed to transition OUT of this tier."),
    ("Notes", None),
]
for col, (h, c) in enumerate(PLAN_HEADERS, start=1):
    header_cell(plan, 4, col, h, comment=c)

geom_sum = (
    f"IF({P('tier_time_growth')}=1,{P('num_tiers')},"
    f"(POWER({P('tier_time_growth')},{P('num_tiers')})-1)/({P('tier_time_growth')}-1))"
)

for i in range(1, 11):
    r = 4 + i
    formula_cell(plan, r, 1, i, number_format="0", font=INPUT_FONT)
    name_row = TIER_NAME_START_ROW + (i - 1)
    formula_cell(plan, r, 2, f"=Parameters!$B${name_row}", font=LINK_FONT,
                 alignment=LEFT_NOWRAP)
    formula_cell(plan, r, 3,
                 f"={P('total_minutes')}*POWER({P('tier_time_growth')},A{r}-1)/{geom_sum}",
                 number_format="0.0")
    formula_cell(plan, r, 4, f"=SUM(C$5:C{r})", number_format="0.0")

    cs = (f"=IF(A{r}<=5,"
          f"{P('click_share_T1')}+({P('click_share_T5_min')}-{P('click_share_T1')})*POWER((A{r}-1)/4,2),"
          f"IF(A{r}<=9,"
          f"{P('click_share_T5_min')}+({P('click_share_T9')}-{P('click_share_T5_min')})*POWER((A{r}-5)/4,2),"
          f"{P('click_share_T10')}))")
    formula_cell(plan, r, 5, cs, number_format="0%")

    formula_cell(plan, r, 6,
                 f"=IF(A{r}<{P('num_tiers')},{P('cohesion_T1_to_T2')}*POWER({P('cohesion_growth')},A{r}-1),0)",
                 number_format="0.0")

    note_text = ""
    if i == 1: note_text = "Tutorial."
    elif i == 2: note_text = "Autoclickers introduced."
    elif i == 5: note_text = "PEAK."
    elif i == 10: note_text = "Endgame."
    nc = plan.cell(row=r, column=7, value=note_text)
    nc.font = NOTE_FONT
    nc.alignment = LEFT_NOWRAP

# Total row
trow = 15
label(plan, trow, 1, "TOTAL")
plan.cell(row=trow, column=1).font = H3
formula_cell(plan, trow, 3, "=SUM(C5:C14)", number_format="0.0")
plan.cell(row=trow, column=3).font = H3
formula_cell(plan, trow, 6, "=SUM(F5:F14)", number_format="0.0")
plan.cell(row=trow, column=6).font = H3

plan.freeze_panes = "B5"


# ============================================================
# UpgradeSim — Tier 1
# ============================================================
sim = wb.create_sheet("UpgradeSim")

# Column layout (v1.1.1 — 7 T1 upgrades + Base contribution column):
# A: Tick                     U: Buy?
# B: Time (min)               V: Cost
# C: Mass start               W: Coh delta
# D: Cohesion                 X: Mass end
# E-K: Levels (SW AB SC Mag OR HP FP)
# L: Transitioned             Y: M/sec rate
# M: MPC                      Z: Note
# N: MPS                      AA-AF: contribution displays
# O: APS                        (SW, AB, SC, Mag, FP, Base)
# P: Click inc/tick           AG-AM: cost helpers (hidden)
# Q: Auto inc/tick            AN: trans available?
# R: Pass inc/tick            AO: could-have-transitioned diagnostic
# S: Total inc/tick           AP: next OS cost
# T: Mass post-inc            AQ: cheapest stackable cost
#                             AR: save mode
#                             AS-AV: stackable scores
#                             AW: action

sim_widths = {
    "A": 6,   "B": 8,   "C": 12,  "D": 8,
    "E": 5,   "F": 5,   "G": 5,   "H": 5,   "I": 5,   "J": 5,   "K": 5,
    "L": 6,
    "M": 10,  "N": 10,  "O": 10,
    "P": 10,  "Q": 10,  "R": 10,  "S": 10,  "T": 12,
    "U": 8,   "V": 10,  "W": 9,   "X": 12,
    "Y": 10,  "Z": 28,
    "AA": 11, "AB": 11, "AC": 11, "AD": 11, "AE": 11, "AF": 11,
    "AG": 9,  "AH": 9,  "AI": 9,  "AJ": 9,  "AK": 9,  "AL": 9,  "AM": 9,
    "AN": 8,  "AO": 8,  "AP": 11, "AQ": 11, "AR": 9,
    "AS": 12, "AT": 12, "AU": 12, "AV": 12,
    "AW": 12, "AX": 9,
}
set_widths(sim, sim_widths)

st = sim.cell(row=1, column=1, value="UpgradeSim — Tier 1 (Solar System)")
st.font = H1
st.alignment = LEFT_NOWRAP

ss = sim.cell(row=2, column=1,
              value="10-second tick simulation. Game ticks every 1 second under the hood; each row aggregates 10 game ticks. Sim runs the COMPLETION path — won't transition until all completionist upgrades are owned.")
ss.font = NOTE_FONT
ss.alignment = LEFT_NOWRAP

HEADER_ROW = 4
SIM_HEADERS = [
    ("Tick", "Sim tick number. Each tick = 10 game seconds."),
    ("Time (min)", "Minutes elapsed from game start."),
    ("Mass start", "Mass at start of tick."),
    ("Cohesion", "Cohesion at start of tick."),
    ("SW", "Solar Wind levels."),
    ("AB", "Asteroid Belt levels."),
    ("SC", "Stellar Coupling levels."),
    ("Mag", "Magnetosphere levels (max 3)."),
    ("OR", "Orbital Resonance owned."),
    ("HP", "Heliopause owned."),
    ("FP", "First Photons owned."),
    ("T2?", "Transitioned to Tier 2."),
    ("MPC", "Mass per Click."),
    ("MPS", "Mass per Second (passive)."),
    ("APS", "Autoclicks per Second. Always 0 in T1."),
    ("Click inc", "Active click income this tick (10s)."),
    ("Auto inc", "Autoclicker income this tick."),
    ("Pass inc", "Passive income this tick."),
    ("Total inc", None),
    ("Mass post-inc", None),
    ("Buy?", None),
    ("Cost", None),
    ("Coh +", None),
    ("Mass end", None),
    ("M/sec", "Implied mass-per-second rate at end of tick."),
    ("Note", None),
]
for col, (h, c) in enumerate(SIM_HEADERS, start=1):
    header_cell(sim, HEADER_ROW, col, h, comment=c)

# Per-upgrade contribution display headers (AA-AF)
CONTRIB_HEADERS = [
    ("SW M/min", "Solar Wind's contribution to total income, in Mass/min. Already includes any global multipliers."),
    ("AB M/min", "Asteroid Belt's contribution in Mass/min. Already includes any global multipliers."),
    ("SC M/min", "Stellar Coupling's contribution to active click income, in Mass/min."),
    ("Mag M/min", "Magnetosphere's contribution to passive income, in Mass/min. Already includes any global multipliers."),
    ("FP M/min", "First Photons' direct M/min contribution (its base 0.5 M/sec). Its click multiplier effect is rolled into SC's column AND the Base column."),
    ("Base M/min", "Income from the implicit base click value (1.0 mass/click), with all global multipliers applied. Includes both active clicks and autoclickers (autoclickers are 0 in T1). Adding all upgrade columns plus this Base column equals total income."),
]
for col, (h, c) in enumerate(CONTRIB_HEADERS, start=27):  # AA=27, AF=32
    header_cell(sim, HEADER_ROW, col, h, comment=c)

# Hidden helper headers (AF onwards)
HELPER_HEADERS = [
    "SW cost", "AB cost", "SC cost", "Mag cost",
    "OR cost", "HP cost", "FP cost",
    "Trans?", "Could-trans?", "Next OS cost", "Cheapest stack", "Save mode",
    "SW VPC", "AB VPC", "SC VPC", "Mag VPC",
    "Next OS VPC",
    "Action",
]
for col, h in enumerate(HELPER_HEADERS, start=33):  # AG=33
    header_cell(sim, HEADER_ROW, col, h)

# Hide helper columns by default (AG onwards, including new AX)
for col_letter in ["AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO",
                   "AP", "AQ", "AR", "AS", "AT", "AU", "AV", "AW", "AX"]:
    sim.column_dimensions[col_letter].hidden = True


def U(name, field_letter):
    """Reference to a cell on Upgrades tab for this upgrade."""
    return f"Upgrades!${field_letter}${UPG_ROW[name]}"


# Sim level columns map. Order matches the level columns E-K on the sim.
LVL_COL = {
    "Solar Wind": "E", "Asteroid Belt": "F", "Stellar Coupling": "G",
    "Magnetosphere": "H",
    "Orbital Resonance": "I", "Heliopause": "J",
    "First Photons": "K",
}
# All upgrades, in level-column order (matters for some helper math)
NAMES_T1 = ["Solar Wind", "Asteroid Belt", "Stellar Coupling", "Magnetosphere",
            "Orbital Resonance", "Heliopause", "First Photons"]
# Stackable upgrades (max_levels > 1)
STACKABLES = ["Solar Wind", "Asteroid Belt", "Stellar Coupling", "Magnetosphere"]
# One-shot upgrades (max_levels = 1)
ONESHOTS = ["Orbital Resonance", "Heliopause", "First Photons"]
# Completionist upgrades (must be owned for transition under completion path)
COMPLETIONIST = ["Magnetosphere", "First Photons"]
SHORT = {"Solar Wind": "SW", "Asteroid Belt": "AB", "Stellar Coupling": "SC",
         "Magnetosphere": "Mag",
         "Orbital Resonance": "OR", "Heliopause": "HP",
         "First Photons": "FP"}


def self_contrib_formula(name, stat, r):
    """One upgrade's self contribution to a stat at sim row r (BEFORE incoming synergies)."""
    field_map = {'mps': ('G', 'H', 'I'), 'mpc': ('J', 'K', 'L'), 'aps': ('M', 'N', 'O')}
    base_col, add_col, self_col = field_map[stat]
    lvl = f"{LVL_COL[name]}{r}"
    return (f"IF({lvl}=0,0,"
            f"({U(name, base_col)}+{lvl}*{U(name, add_col)})*POWER({U(name, self_col)},{lvl}))")


def synergy_mult_formula(target_name, r):
    """Product of incoming synergy multipliers targeting `target_name` at row r.
    For each potential provider P (any other upgrade), if P's synergy_target == target_name,
    apply (P.synergy_mult ^ P.level). Otherwise apply 1. Multiply all factors together."""
    parts = []
    target_str = f'"{target_name}"'
    for provider in NAMES_T1:
        if provider == target_name:
            continue  # an upgrade can't synergize with itself
        provider_target_cell = f"Upgrades!$T${UPG_ROW[provider]}"  # synergy target column
        provider_mult_cell = f"Upgrades!$U${UPG_ROW[provider]}"     # synergy mult column
        provider_level_cell = f"{LVL_COL[provider]}{r}"
        parts.append(
            f'IF({provider_target_cell}={target_str},'
            f'POWER({provider_mult_cell},{provider_level_cell}),1)'
        )
    if not parts:
        return "1"
    return "*".join(parts)


def global_mult_formula(stat, r):
    """Product of (mult_all_<stat> ^ level) across all upgrades."""
    field_map = {'mps': 'P', 'mpc': 'Q', 'aps': 'R'}
    field = field_map[stat]
    parts = [f"POWER({U(n, field)},{LVL_COL[n]}{r})" for n in NAMES_T1]
    return "*".join(parts)


def aggregate_formula(stat, r, base_value=None):
    """Total stat at sim row r, with each upgrade's self-contribution multiplied by its incoming synergy."""
    self_terms = []
    for n in NAMES_T1:
        base = self_contrib_formula(n, stat, r)
        synergy = synergy_mult_formula(n, r)
        self_terms.append(f"({base})*({synergy})")
    summed = "+".join(self_terms)
    if base_value is not None:
        sum_part = f"({base_value}+{summed})"
    else:
        sum_part = f"({summed})"
    return f"={sum_part}*{global_mult_formula(stat, r)}"


# Helper column letter assignments (computed for clarity below)
# Cost columns AG-AM (7 upgrades, cost per upgrade) -- shifted by 1 from v1.1 to make room for Base contribution column at AF
COST_COL = {
    "Solar Wind": "AG", "Asteroid Belt": "AH", "Stellar Coupling": "AI",
    "Magnetosphere": "AJ",
    "Orbital Resonance": "AK", "Heliopause": "AL", "First Photons": "AM",
}
# v1.2.1: VPC columns AS-AV (4 stackables) replace the old "stackable score" columns
# Each holds value-per-cost = (per-level marginal income) / (next purchase cost)
VPC_COL = {"Solar Wind": "AS", "Asteroid Belt": "AT",
           "Stellar Coupling": "AU", "Magnetosphere": "AV"}
TRANS_COL = "AN"
COULD_TRANS_COL = "AO"
NEXT_OS_COL = "AP"
CHEAPEST_STACK_COL = "AQ"
SAVE_MODE_COL = "AR"
NEXT_OS_VPC_COL = "AW"  # NEW: VPC of the cheapest unowned one-shot
ACTION_COL = "AX"       # shifted from AW to AX to make room for NEXT_OS_VPC_COL

SIM_ROWS = 400  # extended for v1.2's higher one-shot costs (HP=575, FP=850, Mag-to-5=2480)
SIM_START = HEADER_ROW + 1  # row 5

# Tick 0 (init row)
init = SIM_START
formula_cell(sim, init, 1, 0, number_format="0")
formula_cell(sim, init, 2, 0, number_format="0.00")
formula_cell(sim, init, 3, 0, number_format="0.0")
formula_cell(sim, init, 4, 0, number_format="0.00")
# Levels E-K (7 upgrades) and T2? (L)
for c in range(5, 13):
    formula_cell(sim, init, c, 0, number_format="0")

# M: MPC = base click value
formula_cell(sim, init, 13,
             aggregate_formula('mpc', init, base_value=P('base_click_value')),
             number_format="0.000", fill=DERIVED_FILL)
# N: MPS, O: APS
formula_cell(sim, init, 14, aggregate_formula('mps', init),
             number_format="0.000", fill=DERIVED_FILL)
formula_cell(sim, init, 15, aggregate_formula('aps', init),
             number_format="0.000", fill=DERIVED_FILL)

# P-S: Income breakdowns (all 0 at tick 0)
for c in range(16, 20):
    formula_cell(sim, init, c, 0, number_format="0.00")
# T-X: Mass post-inc, buy, cost, coh delta, mass end
formula_cell(sim, init, 20, 0, number_format="0.0")
formula_cell(sim, init, 21, "")
formula_cell(sim, init, 22, 0, number_format="0.0")
formula_cell(sim, init, 23, 0, number_format="0.00")
formula_cell(sim, init, 24, 0, number_format="0.0")
# Y: M/sec rate
formula_cell(sim, init, 25,
             f"=M{init}*{P('active_clicks_per_min')}/60*{P('active_play_fraction')}+"
             f"O{init}*M{init}+N{init}",
             number_format="0.00", fill=DERIVED_FILL)
nc = sim.cell(row=init, column=26, value="game start")
nc.font = EVENT_FONT
nc.alignment = LEFT_NOWRAP

# AA-AF: contribution displays at tick 0 — use the same formulas as per-tick rows.
# At tick 0 with all levels = 0, SW/AB/SC/Mag/FP all evaluate to 0.
# Base evaluates to base_click × global_mpc × cpm × active_frac = 22.5 M/min,
# which matches the M/sec rate column showing 0.375 M/sec at tick 0.
global_mps_init = global_mult_formula('mps', init)
global_mpc_init = global_mult_formula('mpc', init)
formula_cell(sim, init, 27,
             f"=({self_contrib_formula('Solar Wind','mps',init)})*({synergy_mult_formula('Solar Wind',init)})*{global_mps_init}*60",
             number_format="0.00")
formula_cell(sim, init, 28,
             f"=({self_contrib_formula('Asteroid Belt','mps',init)})*({synergy_mult_formula('Asteroid Belt',init)})*{global_mps_init}*60",
             number_format="0.00")
formula_cell(sim, init, 29,
             f"=({self_contrib_formula('Stellar Coupling','mpc',init)})*({synergy_mult_formula('Stellar Coupling',init)})*{global_mpc_init}*"
             f"{P('active_clicks_per_min')}*{P('active_play_fraction')}",
             number_format="0.00")
formula_cell(sim, init, 30,
             f"=({self_contrib_formula('Magnetosphere','mps',init)})*({synergy_mult_formula('Magnetosphere',init)})*{global_mps_init}*60",
             number_format="0.00")
formula_cell(sim, init, 31,
             f"=({self_contrib_formula('First Photons','mps',init)})*({synergy_mult_formula('First Photons',init)})*{global_mps_init}*60",
             number_format="0.00")
formula_cell(sim, init, 32,
             f"={P('base_click_value')}*{global_mpc_init}*"
             f"({P('active_clicks_per_min')}*{P('active_play_fraction')}+O{init}*60)",
             number_format="0.00")

# AF-AL: cost helpers at tick 0
def cost_formula(name, r):
    """Cost of next purchase of this upgrade. '' if at max level."""
    lvl = f"{LVL_COL[name]}{r}"
    return (f'=IF({lvl}>={U(name,"E")},"",'
            f'{U(name,"C")}*POWER({U(name,"D")},{lvl}))')

# Stackable costs
for name in ["Solar Wind", "Asteroid Belt", "Stellar Coupling", "Magnetosphere"]:
    col_idx = ord(COST_COL[name][1]) - ord('A') + 1 + 26  # AF=32, AG=33, AH=34, AI=35
    # Use simpler col index calc:
col_letter_to_idx = lambda L: (ord(L[0]) - ord('A') + 1) * 26 + (ord(L[1]) - ord('A') + 1) if len(L) == 2 else ord(L) - ord('A') + 1

for name, col_letter in COST_COL.items():
    if len(col_letter) == 2:
        idx = (ord(col_letter[0]) - ord('A') + 1) * 26 + (ord(col_letter[1]) - ord('A') + 1)
    else:
        idx = ord(col_letter) - ord('A') + 1
    formula_cell(sim, init, idx, cost_formula(name, init), number_format="0.0")

def col_idx(letter):
    """Column letter to 1-based index. Handles 1- and 2-letter columns."""
    if len(letter) == 1:
        return ord(letter) - ord('A') + 1
    return (ord(letter[0]) - ord('A') + 1) * 26 + (ord(letter[1]) - ord('A') + 1)

# Trans? — true when cohesion met AND all completionist owned AND not yet transitioned
completionist_check = "*".join(f"({LVL_COL[n]}{init}>={U(n,'E')})" for n in COMPLETIONIST)
formula_cell(sim, init, col_idx(TRANS_COL),
             f'=IF(AND(D{init}>={P("cohesion_T1_to_T2")},{completionist_check}=1,L{init}=0),1,"")',
             number_format="0")
# Could-trans? — true when cohesion met (regardless of completionist)
formula_cell(sim, init, col_idx(COULD_TRANS_COL),
             f'=IF(AND(D{init}>={P("cohesion_T1_to_T2")},L{init}=0),1,"")',
             number_format="0")

# Next OS cost = MIN over unowned one-shots
oneshot_cost_terms = ",".join(
    f'IF(ISNUMBER({COST_COL[n]}{init}),{COST_COL[n]}{init},9E15)'
    for n in ONESHOTS)
formula_cell(sim, init, col_idx(NEXT_OS_COL),
             f"=MIN({oneshot_cost_terms})", number_format="0.0")

# Cheapest stackable cost
stack_cost_terms = ",".join(
    f'IF(ISNUMBER({COST_COL[n]}{init}),{COST_COL[n]}{init},9E15)'
    for n in STACKABLES)
formula_cell(sim, init, col_idx(CHEAPEST_STACK_COL),
             f"=MIN({stack_cost_terms})", number_format="0.0")

# Save mode at tick 0 — no purchases possible since mass=0, just set FALSE
formula_cell(sim, init, col_idx(SAVE_MODE_COL), "=FALSE")

# Stackable VPCs at tick 0 (everything's at level 0, mass is 0, nothing affordable)
# Set to 0 since computing them serves no purpose at tick 0 with no levels
for name in STACKABLES:
    formula_cell(sim, init, col_idx(VPC_COL[name]), 0,
                 number_format="0.000000")

# Next OS VPC at tick 0 — also 0 (everything is at base state, no useful targets yet)
formula_cell(sim, init, col_idx(NEXT_OS_VPC_COL), 0,
             number_format="0.000000")

# Action at tick 0
formula_cell(sim, init, col_idx(ACTION_COL), "")

# Per-tick rows
for tick in range(1, SIM_ROWS + 1):
    r = SIM_START + tick
    pr = r - 1

    # A: tick
    formula_cell(sim, r, 1, tick, number_format="0")
    # B: time min
    formula_cell(sim, r, 2, f"=A{r}*10/60", number_format="0.00")
    # C: mass start = previous mass end (X)
    formula_cell(sim, r, 3, f"=X{pr}", number_format="0.0")
    # D: cohesion = previous + previous coh delta
    formula_cell(sim, r, 4, f"=D{pr}+W{pr}", number_format="0.00")

    # E-K: levels carry forward + buy increment (using SHORT codes)
    # Stackables (E, F, G, H): allow level increment
    formula_cell(sim, r, 5, f'=E{pr}+IF(U{pr}="SW",1,0)', number_format="0")
    formula_cell(sim, r, 6, f'=F{pr}+IF(U{pr}="AB",1,0)', number_format="0")
    formula_cell(sim, r, 7, f'=G{pr}+IF(U{pr}="SC",1,0)', number_format="0")
    formula_cell(sim, r, 8, f'=H{pr}+IF(U{pr}="Mag",1,0)', number_format="0")
    # One-shots (I, J, K): MAX (cap at 1)
    formula_cell(sim, r, 9, f'=MAX(I{pr},IF(U{pr}="OR",1,0))', number_format="0")
    formula_cell(sim, r, 10, f'=MAX(J{pr},IF(U{pr}="HP",1,0))', number_format="0")
    formula_cell(sim, r, 11, f'=MAX(K{pr},IF(U{pr}="FP",1,0))', number_format="0")
    # L: transitioned flag
    formula_cell(sim, r, 12, f'=MAX(L{pr},IF(U{pr}="transition",1,0))', number_format="0")

    # M: MPC, N: MPS, O: APS
    formula_cell(sim, r, 13, aggregate_formula('mpc', r, base_value=P('base_click_value')),
                 number_format="0.000", fill=DERIVED_FILL)
    formula_cell(sim, r, 14, aggregate_formula('mps', r),
                 number_format="0.000", fill=DERIVED_FILL)
    formula_cell(sim, r, 15, aggregate_formula('aps', r),
                 number_format="0.000", fill=DERIVED_FILL)

    # P: click income/tick (zero after transition)
    formula_cell(sim, r, 16,
                 f"=IF(L{r}=1,0,M{r}*{P('active_clicks_per_min')}/60*{P('active_play_fraction')}*10)",
                 number_format="0.00")
    # Q: autoclick income/tick
    formula_cell(sim, r, 17, f"=IF(L{r}=1,0,O{r}*M{r}*10)", number_format="0.00")
    # R: passive income/tick
    formula_cell(sim, r, 18, f"=IF(L{r}=1,0,N{r}*10)", number_format="0.00")
    # S: total income/tick
    formula_cell(sim, r, 19, f"=P{r}+Q{r}+R{r}", number_format="0.00")
    # T: mass post-income
    formula_cell(sim, r, 20, f"=C{r}+S{r}", number_format="0.0")

    # ---- Helper columns ----
    # AF-AL: costs
    for name in NAMES_T1:
        formula_cell(sim, r, col_idx(COST_COL[name]),
                     cost_formula(name, r), number_format="0.0")

    # AM: trans? — cohesion met AND all completionist at max AND not yet transitioned
    completionist_check_r = "*".join(f"({LVL_COL[n]}{r}>={U(n,'E')})" for n in COMPLETIONIST)
    formula_cell(sim, r, col_idx(TRANS_COL),
                 f'=IF(AND(D{r}>={P("cohesion_T1_to_T2")},{completionist_check_r}=1,L{r}=0),1,"")',
                 number_format="0")
    # AN: could-trans? — cohesion met (regardless of completionist)
    formula_cell(sim, r, col_idx(COULD_TRANS_COL),
                 f'=IF(AND(D{r}>={P("cohesion_T1_to_T2")},L{r}=0),1,"")',
                 number_format="0")

    # AO: next OS cost (cheapest unowned one-shot)
    oneshot_cost_terms_r = ",".join(
        f'IF(ISNUMBER({COST_COL[n]}{r}),{COST_COL[n]}{r},9E15)' for n in ONESHOTS)
    formula_cell(sim, r, col_idx(NEXT_OS_COL),
                 f"=MIN({oneshot_cost_terms_r})", number_format="0.0")
    # AP: cheapest stackable
    stack_cost_terms_r = ",".join(
        f'IF(ISNUMBER({COST_COL[n]}{r}),{COST_COL[n]}{r},9E15)' for n in STACKABLES)
    formula_cell(sim, r, col_idx(CHEAPEST_STACK_COL),
                 f"=MIN({stack_cost_terms_r})", number_format="0.0")
    # AQ: save mode (NEW v1.2.1: VPC-based threshold instead of static save_ratio)
    # save when next_os_vpc > 1.5 × max affordable stackable VPC
    # We need this AFTER computing VPCs, so just reference the columns by letter
    # (Excel handles the dependency order automatically)
    #
    # v1.2.1 post-cohesion focus: once cohesion is met (AO=1) but completionist is
    # not yet maxed (AN<>1), non-completionist stackables (SW, AB, SC) are GATED OFF.
    # This matches observed player behavior — once they see the finish line, they
    # stop buying SW/AB/SC and rush completionist (Mag) completion. The completionist
    # stackable (Mag) is the only stackable still considered affordable in this phase.
    post_coh_focus = f'AND({COULD_TRANS_COL}{r}=1,{TRANS_COL}{r}<>1)'
    sw_aff = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL["Solar Wind"]}{r}),{COST_COL["Solar Wind"]}{r}<=T{r}),{VPC_COL["Solar Wind"]}{r},0)'
    ab_aff = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL["Asteroid Belt"]}{r}),{COST_COL["Asteroid Belt"]}{r}<=T{r}),{VPC_COL["Asteroid Belt"]}{r},0)'
    sc_aff = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL["Stellar Coupling"]}{r}),{COST_COL["Stellar Coupling"]}{r}<=T{r}),{VPC_COL["Stellar Coupling"]}{r},0)'
    mag_aff = f'IF(AND(ISNUMBER({COST_COL["Magnetosphere"]}{r}),{COST_COL["Magnetosphere"]}{r}<=T{r}),{VPC_COL["Magnetosphere"]}{r},0)'
    max_aff_vpc_expr = f"MAX({sw_aff},{ab_aff},{sc_aff},{mag_aff})"
    formula_cell(sim, r, col_idx(SAVE_MODE_COL),
                 f"=IF({NEXT_OS_VPC_COL}{r}>{P('save_vpc_threshold')}*({max_aff_vpc_expr}),TRUE,FALSE)")

    # AS-AV: stackable VPCs (per-level marginal income / next purchase cost)
    # SW (passive): +M/sec/lvl × global_mps × incoming_synergy / cost
    # AB (passive): same form
    # Mag (passive): same form
    # SC (click):   +M/click/lvl × global_mpc × incoming_synergy × cpm × active_frac / 60 / cost
    for name in ["Solar Wind", "Asteroid Belt", "Magnetosphere"]:
        delta = (f"{U(name,'H')}*({global_mult_formula('mps', r)})"
                 f"*({synergy_mult_formula(name, r)})")
        formula_cell(sim, r, col_idx(VPC_COL[name]),
                     f'=IFERROR({delta}/{COST_COL[name]}{r},0)',
                     number_format="0.000000")
    # SC is click-based
    sc_delta = (f"{U('Stellar Coupling','K')}*({global_mult_formula('mpc', r)})"
                f"*({synergy_mult_formula('Stellar Coupling', r)})"
                f"*{P('active_clicks_per_min')}*{P('active_play_fraction')}/60")
    formula_cell(sim, r, col_idx(VPC_COL["Stellar Coupling"]),
                 f'=IFERROR({sc_delta}/{COST_COL["Stellar Coupling"]}{r},0)',
                 number_format="0.000000")

    # AW: SAVE TARGET VPC = max VPC across all UNAFFORDABLE buyable upgrades
    # This includes:
    #   - Unowned one-shots (OR, HP, FP) — by definition unaffordable when not bought yet
    #   - Stackables that are currently unaffordable (especially Mag whose costs balloon)
    # The "save mode" check then asks: is anything I'm waiting for worth more than what I can afford now?

    # OR: ×1.25 to all M/sec → delta = 0.25 × current MPS (col N)
    or_vpc = f'IF(I{r}>0,0,IFERROR(0.25*N{r}/{U("Orbital Resonance","C")},0))'
    # HP: ×1.5 synergy on SC's self-contrib → delta = 0.5 × SC's contribution to M/sec
    sc_self_mpc_r = self_contrib_formula("Stellar Coupling", "mpc", r)
    sc_synergy_r = synergy_mult_formula("Stellar Coupling", r)
    sc_per_sec_contrib = (f"({sc_self_mpc_r})*({sc_synergy_r})"
                         f"*({global_mult_formula('mpc', r)})"
                         f"*{P('active_clicks_per_min')}*{P('active_play_fraction')}/60")
    hp_vpc = f'IF(J{r}>0,0,IFERROR(0.5*({sc_per_sec_contrib})/{U("Heliopause","C")},0))'
    # FP: +1.0 base M/sec + ×1.20 to all M/click
    fp_delta_passive = f"1.0*({global_mult_formula('mps', r)})"
    fp_delta_click = f"0.20*M{r}*{P('active_clicks_per_min')}*{P('active_play_fraction')}/60"
    fp_vpc = f'IF(K{r}>0,0,IFERROR(({fp_delta_passive}+{fp_delta_click})/{U("First Photons","C")},0))'

    # Stackable save VPCs: each stackable's VPC IF unaffordable (otherwise it'd be in max_aff_vpc)
    # Use the already-computed VPC columns AS-AV
    sw_save = f'IF(AND(ISNUMBER({COST_COL["Solar Wind"]}{r}),{COST_COL["Solar Wind"]}{r}>T{r}),{VPC_COL["Solar Wind"]}{r},0)'
    ab_save = f'IF(AND(ISNUMBER({COST_COL["Asteroid Belt"]}{r}),{COST_COL["Asteroid Belt"]}{r}>T{r}),{VPC_COL["Asteroid Belt"]}{r},0)'
    sc_save = f'IF(AND(ISNUMBER({COST_COL["Stellar Coupling"]}{r}),{COST_COL["Stellar Coupling"]}{r}>T{r}),{VPC_COL["Stellar Coupling"]}{r},0)'
    mag_save = f'IF(AND(ISNUMBER({COST_COL["Magnetosphere"]}{r}),{COST_COL["Magnetosphere"]}{r}>T{r}),{VPC_COL["Magnetosphere"]}{r},0)'

    formula_cell(sim, r, col_idx(NEXT_OS_VPC_COL),
                 f"=MAX({or_vpc},{hp_vpc},{fp_vpc},{sw_save},{ab_save},{sc_save},{mag_save})",
                 number_format="0.000000")

    # AX: action (VPC-based)
    # Priority:
    #   1. transition if available
    #   2. buy cheapest affordable one-shot (cohesion priority preserved)
    #   3. if save mode (VPC threshold met), do nothing
    #   4. else buy stackable with max VPC (from affordable)
    affordable_OR = f'AND(I{r}=0,ISNUMBER({COST_COL["Orbital Resonance"]}{r}),{COST_COL["Orbital Resonance"]}{r}<=T{r})'
    affordable_HP = f'AND(J{r}=0,ISNUMBER({COST_COL["Heliopause"]}{r}),{COST_COL["Heliopause"]}{r}<=T{r})'
    affordable_FP = f'AND(K{r}=0,ISNUMBER({COST_COL["First Photons"]}{r}),{COST_COL["First Photons"]}{r}<=T{r})'
    cheapest_affordable_os = (
        f"MIN("
        f"IF({affordable_OR},{COST_COL['Orbital Resonance']}{r},9E15),"
        f"IF({affordable_HP},{COST_COL['Heliopause']}{r},9E15),"
        f"IF({affordable_FP},{COST_COL['First Photons']}{r},9E15)"
        f")")
    pick_OR = f'AND({affordable_OR},{COST_COL["Orbital Resonance"]}{r}={cheapest_affordable_os})'
    pick_HP = f'AND({affordable_HP},{COST_COL["Heliopause"]}{r}={cheapest_affordable_os})'

    # For stackable selection: pick the one with max VPC among affordable
    # max_aff_vpc_expr already defined above for save mode
    # We need to also ensure max_aff_vpc > 0 (otherwise nothing affordable to buy)
    # Match the stackable that has VPC == max_aff_vpc AND is affordable
    pick_SW = f'AND({sw_aff}>0,{sw_aff}={max_aff_vpc_expr})'
    pick_AB = f'AND({ab_aff}>0,{ab_aff}={max_aff_vpc_expr})'
    pick_SC = f'AND({sc_aff}>0,{sc_aff}={max_aff_vpc_expr})'
    pick_Mag = f'AND({mag_aff}>0,{mag_aff}={max_aff_vpc_expr})'

    formula_cell(sim, r, col_idx(ACTION_COL),
                 f'=IF({TRANS_COL}{r}=1,"transition",'
                 f'IF({cheapest_affordable_os}<9E15,'
                   f'IF({pick_OR},"OR",IF({pick_HP},"HP","FP")),'
                 f'IF({SAVE_MODE_COL}{r}=TRUE,"",'
                 f'IF({max_aff_vpc_expr}<=0,"",'
                 f'IF({pick_SW},"SW",'
                 f'IF({pick_AB},"AB",'
                 f'IF({pick_SC},"SC","Mag")))))))')

    # U: Buy? = action
    formula_cell(sim, r, 21, f"={ACTION_COL}{r}", font=EVENT_FONT, alignment=CENTER)
    # V: Cost
    formula_cell(sim, r, 22,
                 f'=IF(U{r}="",0,IF(U{r}="transition",0,'
                 f'IF(U{r}="SW",{COST_COL["Solar Wind"]}{r},'
                 f'IF(U{r}="AB",{COST_COL["Asteroid Belt"]}{r},'
                 f'IF(U{r}="SC",{COST_COL["Stellar Coupling"]}{r},'
                 f'IF(U{r}="Mag",{COST_COL["Magnetosphere"]}{r},'
                 f'IF(U{r}="OR",{COST_COL["Orbital Resonance"]}{r},'
                 f'IF(U{r}="HP",{COST_COL["Heliopause"]}{r},'
                 f'IF(U{r}="FP",{COST_COL["First Photons"]}{r},0)))))))))',
                 number_format="0.0")
    # W: Cohesion delta — look up cohesion_reward from Upgrades by upgrade name
    formula_cell(sim, r, 23,
                 f'=IF(U{r}="OR",{U("Orbital Resonance","F")},'
                 f'IF(U{r}="HP",{U("Heliopause","F")},'
                 f'IF(U{r}="transition",-{P("cohesion_T1_to_T2")},0)))',
                 number_format="0.00")
    # X: mass end
    formula_cell(sim, r, 24, f"=T{r}-V{r}", number_format="0.0")
    # Y: M/sec rate (steady, even after transition)
    formula_cell(sim, r, 25,
                 f"=M{r}*{P('active_clicks_per_min')}/60*{P('active_play_fraction')}+"
                 f"O{r}*M{r}+N{r}",
                 number_format="0.00", fill=DERIVED_FILL)
    # Z: note
    formula_cell(sim, r, 26,
                 f'=IF(U{r}="transition","→ Tier 2 (Stellar Neighborhood)",'
                 f'IF(U{r}="FP","First Photons — light begins, system quickens",'
                 f'IF(U{r}="HP","Heliopause acquired",'
                 f'IF(U{r}="OR","Orbital Resonance — passive multiplier active",'
                 f'IF(AND({COULD_TRANS_COL}{r}=1,L{r}=0,{TRANS_COL}{r}<>1),'
                   f'"Cohesion ready (waiting on completionist upgrades)","")))))',
                 font=EVENT_FONT, alignment=LEFT_NOWRAP)

    # AA-AE: per-upgrade contribution displays in M/min
    global_mps = global_mult_formula('mps', r)
    global_mpc = global_mult_formula('mpc', r)

    # SW
    formula_cell(sim, r, 27,
                 f"=({self_contrib_formula('Solar Wind','mps',r)})*({synergy_mult_formula('Solar Wind',r)})*{global_mps}*60",
                 number_format="0.00")
    # AB
    formula_cell(sim, r, 28,
                 f"=({self_contrib_formula('Asteroid Belt','mps',r)})*({synergy_mult_formula('Asteroid Belt',r)})*{global_mps}*60",
                 number_format="0.00")
    # SC
    formula_cell(sim, r, 29,
                 f"=({self_contrib_formula('Stellar Coupling','mpc',r)})*({synergy_mult_formula('Stellar Coupling',r)})*{global_mpc}*"
                 f"{P('active_clicks_per_min')}*{P('active_play_fraction')}",
                 number_format="0.00")
    # Magnetosphere
    formula_cell(sim, r, 30,
                 f"=({self_contrib_formula('Magnetosphere','mps',r)})*({synergy_mult_formula('Magnetosphere',r)})*{global_mps}*60",
                 number_format="0.00")
    # First Photons (its base 0.5 M/sec contribution)
    formula_cell(sim, r, 31,
                 f"=({self_contrib_formula('First Photons','mps',r)})*({synergy_mult_formula('First Photons',r)})*{global_mps}*60",
                 number_format="0.00")
    # Base M/min — income from the implicit base click value (1.0 mass/click) with all
    # global click multipliers applied. Captures both active clicks and autoclicker clicks
    # at base value. The active clicks part = base × global_mpc × cpm × active_frac × 60/60.
    # The autoclicker part = base × global_mpc × APS × 60. APS=0 in T1.
    formula_cell(sim, r, 32,
                 f"={P('base_click_value')}*{global_mpc}*"
                 f"({P('active_clicks_per_min')}*{P('active_play_fraction')}+O{r}*60)",
                 number_format="0.00")

# Headline cells above the data — show speedrun vs completion timings
sim["A2"].value = "10-second tick simulation. Game ticks every 1 second under the hood; sim runs the COMPLETION path. See row 3 for headline timing metrics."
hl1 = sim.cell(row=3, column=1, value="Earliest possible transition (cohesion ready):")
hl1.font = NOTE_FONT
hl1.alignment = LEFT_NOWRAP
sim.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
hl2 = sim.cell(row=3, column=5,
               value=f'=IFERROR(INDEX(B:B,MATCH(1,{COULD_TRANS_COL}:{COULD_TRANS_COL},0))&" min","—")')
hl2.font = LINK_FONT
hl2.alignment = LEFT_NOWRAP

hl3 = sim.cell(row=3, column=8, value="Actual transition (completion path):")
hl3.font = NOTE_FONT
hl3.alignment = LEFT_NOWRAP
sim.merge_cells(start_row=3, start_column=8, end_row=3, end_column=12)
hl4 = sim.cell(row=3, column=13,
               value=f'=IFERROR(INDEX(B:B,MATCH("transition",U:U,0))&" min","—")')
hl4.font = LINK_FONT
hl4.alignment = LEFT_NOWRAP

sim.freeze_panes = "C5"


# ============================================================
# UpgradeSimT2 — Tier 2 (Stellar Neighborhood)
# ============================================================
# Architecture (Option B): a self-contained sheet parallel to UpgradeSim.
# Carryover from T1 lives in a header block at the top (rows 4-7) and is
# referenced by every per-tick row. T1 upgrades are NOT releveled here —
# their effective contribution becomes a "carry MPS/MPC/APS floor" plus
# carry "× all" multipliers that apply to T2-derived contributions too.
#
# Key formulas adapted for T2:
#   total_MPS = (sum of T2 self-contribs × T2 incoming synergies) × global_T2_mps × carry_all_mps + carry_mps_floor
#   total_MPC = (base_click + sum of T2 mpc-contribs × synergies) × global_T2_mpc × carry_all_mpc + carry_mpc_floor (NB)
#   total_APS = (sum of T2 aps-contribs × synergies) × global_T2_aps × carry_all_aps + carry_aps_floor
#
# (NB) carry_mpc_floor is a "M/click floor", but income aggregation uses MPC × cpm,
#      so the carry floor is added at the MPC level, not the income level. That keeps
#      the carry M/click consistent with autoclicker math (autoclicker × MPC includes
#      the carry floor).
#
# Income per tick (10s):
#   click_inc = MPC × cpm/60 × active_frac × 10
#   auto_inc  = APS × MPC × 10
#   pass_inc  = MPS × 10
sim2 = wb.create_sheet("UpgradeSimT2")

# Column layout (parallel to T1's sim, extended for 9 T2 upgrades):
# A: Tick                     U: Buy?
# B: Time (min)               V: Cost
# C: Mass start               W: Coh delta
# D: Cohesion                 X: Mass end
# E-M: Levels (SK LB ML RLO BD BiP PV OC MG)  -- 9 upgrades
# N: Transitioned (T3)        Y: M/sec rate
# O: MPC                      Z: Note
# P: MPS                      AA-AJ: contribution displays per upgrade + base
# Q: APS                      AK-AS: cost helpers (hidden)
# R: Click inc/tick           AT: trans available?
# S: Auto inc/tick            AU: could-trans diagnostic
# T: Total inc/tick (P+R+S)   AV: next OS cost
#                             AW: cheapest stackable cost
#                             AX: save mode
#                             AY-BC: stackable VPCs (5)
#                             BD: next OS save VPC
#                             BE: action

sim2_widths = {
    "A": 6,   "B": 8,   "C": 14,  "D": 8,
    # Levels E-M
    "E": 5, "F": 5, "G": 5, "H": 5, "I": 5,
    "J": 5, "K": 5, "L": 5, "M": 5,
    "N": 6,
    "O": 10, "P": 10, "Q": 10,
    "R": 10, "S": 10, "T": 10,
    "U": 8, "V": 12, "W": 9, "X": 14,
    "Y": 10, "Z": 32,
    # Contributions AA-AJ (9 upgrades + base)
    "AA": 11, "AB": 11, "AC": 11, "AD": 11, "AE": 11,
    "AF": 11, "AG": 11, "AH": 11, "AI": 11, "AJ": 11,
    # Cost helpers AK-AS (9 upgrades)
    "AK": 9, "AL": 9, "AM": 9, "AN": 9, "AO": 9,
    "AP": 9, "AQ": 9, "AR": 9, "AS": 9,
    "AT": 8, "AU": 8, "AV": 11, "AW": 11, "AX": 9,
    # VPCs AY-BC (5 stackables)
    "AY": 12, "AZ": 12, "BA": 12, "BB": 12, "BC": 12,
    "BD": 12, "BE": 12,
}
set_widths(sim2, sim2_widths)

st2 = sim2.cell(row=1, column=1, value="UpgradeSimT2 — Tier 2 (Stellar Neighborhood)")
st2.font = H1
st2.alignment = LEFT_NOWRAP
ss2 = sim2.cell(row=2, column=1,
                value="10-second tick simulation, parallel to UpgradeSim. Carryover state from T1 is shown in the blue block at the top. Mode toggle on Parameters chooses between Completion-exit and Threshold-exit handoff.")
ss2.font = NOTE_FONT
ss2.alignment = LEFT_NOWRAP

# Carryover header block — rows 4-7
# Row 4: labels for what carries
# Row 5: live values (computed from Parameters based on t2_carry_mode)
T2_HEADER_ROW = 4
T2_CARRY_ROW = 5  # live values computed here
# Field layout in carryover block:
# A5: "Carryover" label
# B5..H5: mass | mps_floor | mpc_floor | aps_floor | all_mps | all_mpc | all_aps
carry_labels = [
    ("A4", "Carryover from T1"),
]
sim2.cell(row=4, column=1, value="Carryover from T1").font = H3
sim2.cell(row=4, column=1).fill = SECTION_FILL
sim2.cell(row=4, column=1).alignment = LEFT_NOWRAP
sim2.merge_cells(start_row=4, start_column=1, end_row=4, end_column=8)

# Row 5 sub-headers
sub_headers_t2 = [
    ("A", "Mode"),
    ("B", "Mass"),
    ("C", "MPS floor"),
    ("D", "MPC floor"),
    ("E", "APS floor"),
    ("F", "× all M/s"),
    ("G", "× all M/c"),
    ("H", "× all AC/s"),
]
for col_letter, text in sub_headers_t2:
    c = sim2.cell(row=5, column=col_idx(col_letter), value=text)
    c.font = H3
    c.alignment = CENTER
    c.fill = SECTION_FILL
    c.border = BORDER_ALL

# Row 6 = live carryover values, picked by t2_carry_mode (1 = Completion, 0 = Threshold)
T2_LIVE_ROW = 6
def carry_pick(p_completion, p_threshold):
    return f'=IF({P("t2_carry_mode")}=1,{P(p_completion)},{P(p_threshold)})'

formula_cell(sim2, T2_LIVE_ROW, col_idx("A"),
             f'=IF({P("t2_carry_mode")}=1,"Completion","Threshold")',
             number_format=None, fill=CARRY_FILL, alignment=CENTER)
formula_cell(sim2, T2_LIVE_ROW, col_idx("B"),
             carry_pick("t2_carry_mass_completion", "t2_carry_mass_threshold"),
             number_format="0.0", fill=CARRY_FILL)
formula_cell(sim2, T2_LIVE_ROW, col_idx("C"),
             carry_pick("t2_carry_mps_completion", "t2_carry_mps_threshold"),
             number_format="0.000", fill=CARRY_FILL)
formula_cell(sim2, T2_LIVE_ROW, col_idx("D"),
             carry_pick("t2_carry_mpc_completion", "t2_carry_mpc_threshold"),
             number_format="0.000", fill=CARRY_FILL)
formula_cell(sim2, T2_LIVE_ROW, col_idx("E"),
             f'={P("t2_carry_aps")}',
             number_format="0.000", fill=CARRY_FILL)
formula_cell(sim2, T2_LIVE_ROW, col_idx("F"),
             carry_pick("t2_carry_all_mps_completion", "t2_carry_all_mps_threshold"),
             number_format="0.000", fill=CARRY_FILL)
formula_cell(sim2, T2_LIVE_ROW, col_idx("G"),
             carry_pick("t2_carry_all_mpc_completion", "t2_carry_all_mpc_threshold"),
             number_format="0.000", fill=CARRY_FILL)
formula_cell(sim2, T2_LIVE_ROW, col_idx("H"),
             f'={P("t2_carry_all_aps")}',
             number_format="0.000", fill=CARRY_FILL)

# Cell references for carryover values (used in per-tick formulas)
T2C = {
    "mass":    f"$B${T2_LIVE_ROW}",
    "mps":     f"$C${T2_LIVE_ROW}",
    "mpc":     f"$D${T2_LIVE_ROW}",
    "aps":     f"$E${T2_LIVE_ROW}",
    "all_mps": f"$F${T2_LIVE_ROW}",
    "all_mpc": f"$G${T2_LIVE_ROW}",
    "all_aps": f"$H${T2_LIVE_ROW}",
}

# Row 8 = headline metrics (filled in below after sim rows)
T2_HEADLINE_ROW = 8

# Row 10 = column headers, row 11 = init tick (tick 0), rows 12+ = per-tick rows
T2_HDR_ROW = 10
SIM2_HEADERS = [
    ("Tick", "Sim tick number. Each tick = 10 game seconds."),
    ("Time (min)", "Minutes elapsed in T2 (does NOT include T1 time)."),
    ("Mass start", "Mass at start of tick."),
    ("Cohesion", "Cohesion accumulated within T2 (towards 2.5 for T3 transition)."),
    ("SK", "Stellar Kinematics levels."),
    ("LB", "Local Bubble levels."),
    ("ML", "Microlensing levels."),
    ("RLO", "Roche Lobe Overflow levels (first APS source)."),
    ("BD", "Brown Dwarf levels (max 5)."),
    ("BiP", "Binary Partner owned."),
    ("PV", "Peculiar Velocity owned."),
    ("OC", "Open Cluster owned."),
    ("MG", "Moving Group owned."),
    ("T3?", "Transitioned to Tier 3."),
    ("MPC", "Mass per Click (T2 + carry)."),
    ("MPS", "Mass per Second passive (T2 + carry)."),
    ("APS", "Autoclicks per Second (T2 + carry)."),
    ("Click inc", "Active click income this tick (10s)."),
    ("Auto inc", "Autoclicker income this tick (APS × MPC × 10)."),
    ("Total inc", "Click + auto + passive income this tick."),
    ("Buy?", "Action chosen this tick."),
    ("Cost", "Mass spent this tick."),
    ("Coh +", "Cohesion gained this tick."),
    ("Mass end", "Mass at end of tick."),
    ("M/sec", "Implied mass-per-second rate at end of tick."),
    ("Note", "Annotations for milestone events."),
]
for col, (h, c) in enumerate(SIM2_HEADERS, start=1):
    header_cell(sim2, T2_HDR_ROW, col, h, comment=c)

# Per-upgrade contribution headers (AA-AJ)
CONTRIB_HEADERS_T2 = [
    ("SK M/min", "Stellar Kinematics' M/min (passive)."),
    ("LB M/min", "Local Bubble's M/min (passive). Includes synergy from RLO."),
    ("ML M/min", "Microlensing's M/min (active click contribution). Includes synergy from BiP."),
    ("RLO M/min", "Roche Lobe Overflow's M/min from autoclicks (APS × MPC). Includes synergy from BD."),
    ("BD M/min", "Brown Dwarf's M/min (passive)."),
    ("BiP M/min", "Binary Partner's direct contribution (zero — its only effect is the synergy on ML)."),
    ("PV M/min", "Peculiar Velocity's direct contribution (zero — it's a global multiplier on all MPS)."),
    ("OC M/min", "Open Cluster's direct contribution (zero — its purpose is cohesion gating for T3)."),
    ("MG M/min", "Moving Group's direct M/min (its base 6.0 M/sec)."),
    ("Carry M/min", "Carryover income floor from T1 (passive only — does NOT include click contribution, which flows through MPC)."),
]
for col, (h, c) in enumerate(CONTRIB_HEADERS_T2, start=27):  # AA = 27
    header_cell(sim2, T2_HDR_ROW, col, h, comment=c)

# Hidden helper headers (AK onwards) — costs (9), trans flags (2), next-OS, cheapest-stack, save-mode, VPCs (5), next-OS-save-VPC, action
HELPER_HEADERS_T2 = [
    "SK cost", "LB cost", "ML cost", "RLO cost", "BD cost",
    "BiP cost", "PV cost", "OC cost", "MG cost",
    "Trans?", "Could-trans?", "Next OS cost", "Cheapest stack", "Save mode",
    "SK VPC", "LB VPC", "ML VPC", "RLO VPC", "BD VPC",
    "Next OS VPC",
    "Action",
]
# AK = 37, so headers start at column 37
for col, h in enumerate(HELPER_HEADERS_T2, start=37):
    header_cell(sim2, T2_HDR_ROW, col, h)

# Hide helper columns
for col_letter in ["AK", "AL", "AM", "AN", "AO", "AP", "AQ", "AR", "AS",
                   "AT", "AU", "AV", "AW", "AX",
                   "AY", "AZ", "BA", "BB", "BC", "BD", "BE"]:
    sim2.column_dimensions[col_letter].hidden = True

# T2 upgrade name lists & column maps
NAMES_T2 = ["Stellar Kinematics", "Local Bubble", "Microlensing",
            "Roche Lobe Overflow", "Brown Dwarf",
            "Binary Partner", "Peculiar Velocity", "Open Cluster", "Moving Group"]
STACKABLES_T2 = ["Stellar Kinematics", "Local Bubble", "Microlensing",
                 "Roche Lobe Overflow", "Brown Dwarf"]
ONESHOTS_T2 = ["Binary Partner", "Peculiar Velocity", "Open Cluster", "Moving Group"]
COMPLETIONIST_T2 = ["Brown Dwarf", "Moving Group"]
SHORT_T2 = {"Stellar Kinematics": "SK", "Local Bubble": "LB", "Microlensing": "ML",
            "Roche Lobe Overflow": "RLO", "Brown Dwarf": "BD",
            "Binary Partner": "BiP", "Peculiar Velocity": "PV",
            "Open Cluster": "OC", "Moving Group": "MG"}
# Level columns E-M
LVL_COL_T2 = {"Stellar Kinematics": "E", "Local Bubble": "F", "Microlensing": "G",
              "Roche Lobe Overflow": "H", "Brown Dwarf": "I",
              "Binary Partner": "J", "Peculiar Velocity": "K",
              "Open Cluster": "L", "Moving Group": "M"}
# Cost helper columns AK-AS
COST_COL_T2 = {"Stellar Kinematics": "AK", "Local Bubble": "AL", "Microlensing": "AM",
               "Roche Lobe Overflow": "AN", "Brown Dwarf": "AO",
               "Binary Partner": "AP", "Peculiar Velocity": "AQ",
               "Open Cluster": "AR", "Moving Group": "AS"}
# VPC columns AY-BC (5 stackables)
VPC_COL_T2 = {"Stellar Kinematics": "AY", "Local Bubble": "AZ", "Microlensing": "BA",
              "Roche Lobe Overflow": "BB", "Brown Dwarf": "BC"}
TRANS_COL_T2 = "AT"
COULD_TRANS_COL_T2 = "AU"
NEXT_OS_COL_T2 = "AV"
CHEAPEST_STACK_COL_T2 = "AW"
SAVE_MODE_COL_T2 = "AX"
NEXT_OS_VPC_COL_T2 = "BD"
ACTION_COL_T2 = "BE"


def self_contrib_formula_T2(name, stat, r):
    """T2-specific self-contrib (uses LVL_COL_T2)."""
    field_map = {'mps': ('G', 'H', 'I'), 'mpc': ('J', 'K', 'L'), 'aps': ('M', 'N', 'O')}
    base_col, add_col, self_col = field_map[stat]
    lvl = f"{LVL_COL_T2[name]}{r}"
    return (f"IF({lvl}=0,0,"
            f"({U(name, base_col)}+{lvl}*{U(name, add_col)})*POWER({U(name, self_col)},{lvl}))")


def synergy_mult_formula_T2(target_name, r):
    """Product of incoming synergy multipliers from T2 providers targeting `target_name`."""
    parts = []
    target_str = f'"{target_name}"'
    for provider in NAMES_T2:
        if provider == target_name:
            continue
        provider_target_cell = f"Upgrades!$T${UPG_ROW[provider]}"
        provider_mult_cell = f"Upgrades!$U${UPG_ROW[provider]}"
        provider_level_cell = f"{LVL_COL_T2[provider]}{r}"
        parts.append(
            f'IF({provider_target_cell}={target_str},'
            f'POWER({provider_mult_cell},{provider_level_cell}),1)'
        )
    if not parts:
        return "1"
    return "*".join(parts)


def global_mult_formula_T2(stat, r):
    """Product of (mult_all_<stat> ^ level) across T2 upgrades only."""
    field_map = {'mps': 'P', 'mpc': 'Q', 'aps': 'R'}
    field = field_map[stat]
    parts = [f"POWER({U(n, field)},{LVL_COL_T2[n]}{r})" for n in NAMES_T2]
    return "*".join(parts)


def aggregate_formula_T2(stat, r, base_value=None):
    """T2 stat aggregation = (T2 sum × T2 globals × carry_all_<stat>) + carry_<stat>_floor.
    For mpc, base_value (base_click_value) is included in the T2 sum like T1.
    For mps and aps, base_value is omitted."""
    self_terms = []
    for n in NAMES_T2:
        base = self_contrib_formula_T2(n, stat, r)
        synergy = synergy_mult_formula_T2(n, r)
        self_terms.append(f"({base})*({synergy})")
    summed = "+".join(self_terms)
    if base_value is not None:
        sum_part = f"({base_value}+{summed})"
    else:
        sum_part = f"({summed})"
    carry_all = T2C['all_' + stat]
    carry_floor = T2C[stat]
    # MPC carry is a floor on M/click; MPS/APS likewise
    return f"={sum_part}*{global_mult_formula_T2(stat, r)}*{carry_all}+{carry_floor}"


def cost_formula_T2(name, r):
    lvl = f"{LVL_COL_T2[name]}{r}"
    return (f'=IF({lvl}>={U(name,"E")},"",'
            f'{U(name,"C")}*POWER({U(name,"D")},{lvl}))')


SIM2_ROWS = 400
SIM2_START = T2_HDR_ROW + 1  # row 11 = init (tick 0)

# Tick 0 (init row)
init2 = SIM2_START
formula_cell(sim2, init2, 1, 0, number_format="0")
formula_cell(sim2, init2, 2, 0, number_format="0.00")
# Mass start = carry mass
formula_cell(sim2, init2, 3, f"={T2C['mass']}", number_format="0.0")
# Cohesion = 0 (T2's cohesion accumulator)
formula_cell(sim2, init2, 4, 0, number_format="0.00")
# Levels E-M (9 upgrades) and T3? (N)
for c in range(5, 15):
    formula_cell(sim2, init2, c, 0, number_format="0")

# O: MPC, P: MPS, Q: APS at tick 0
formula_cell(sim2, init2, col_idx("O"),
             aggregate_formula_T2('mpc', init2, base_value=P('base_click_value')),
             number_format="0.000", fill=DERIVED_FILL)
formula_cell(sim2, init2, col_idx("P"),
             aggregate_formula_T2('mps', init2),
             number_format="0.000", fill=DERIVED_FILL)
formula_cell(sim2, init2, col_idx("Q"),
             aggregate_formula_T2('aps', init2),
             number_format="0.000", fill=DERIVED_FILL)

# R-T: incomes (all 0 at tick 0)
for c in [col_idx("R"), col_idx("S"), col_idx("T")]:
    formula_cell(sim2, init2, c, 0, number_format="0.00")
# U-X: action / cost / coh / mass-end
formula_cell(sim2, init2, col_idx("U"), "")
formula_cell(sim2, init2, col_idx("V"), 0, number_format="0.0")
formula_cell(sim2, init2, col_idx("W"), 0, number_format="0.00")
formula_cell(sim2, init2, col_idx("X"), f"=C{init2}", number_format="0.0")
# Y: M/sec rate
formula_cell(sim2, init2, col_idx("Y"),
             f"=O{init2}*{P('active_clicks_per_min')}/60*{P('active_play_fraction')}+"
             f"Q{init2}*O{init2}+P{init2}",
             number_format="0.00", fill=DERIVED_FILL)
nc2 = sim2.cell(row=init2, column=col_idx("Z"), value="T2 start (T1 carryover loaded)")
nc2.font = EVENT_FONT
nc2.alignment = LEFT_NOWRAP

# Contribution displays at tick 0 (all zero except Carry which shows the carry MPS in M/min)
for col_letter in ["AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI"]:
    formula_cell(sim2, init2, col_idx(col_letter), 0, number_format="0.00")
# Carry M/min display
formula_cell(sim2, init2, col_idx("AJ"),
             f"={T2C['mps']}*60", number_format="0.00")

# Cost helpers at tick 0
for name in NAMES_T2:
    formula_cell(sim2, init2, col_idx(COST_COL_T2[name]),
                 cost_formula_T2(name, init2), number_format="0.0")

# Trans? — cohesion 2.5 met AND all completionist owned
completionist_check_t2_init = "*".join(f"({LVL_COL_T2[n]}{init2}>={U(n,'E')})" for n in COMPLETIONIST_T2)
t2_to_t3_cohesion = f"{P('cohesion_T1_to_T2')}*{P('cohesion_growth')}"  # 1.0 × 2.5 = 2.5
formula_cell(sim2, init2, col_idx(TRANS_COL_T2),
             f'=IF(AND(D{init2}>={t2_to_t3_cohesion},{completionist_check_t2_init}=1,N{init2}=0),1,"")',
             number_format="0")
formula_cell(sim2, init2, col_idx(COULD_TRANS_COL_T2),
             f'=IF(AND(D{init2}>={t2_to_t3_cohesion},N{init2}=0),1,"")',
             number_format="0")

oneshot_cost_terms_t2_init = ",".join(
    f'IF(ISNUMBER({COST_COL_T2[n]}{init2}),{COST_COL_T2[n]}{init2},9E15)'
    for n in ONESHOTS_T2)
formula_cell(sim2, init2, col_idx(NEXT_OS_COL_T2),
             f"=MIN({oneshot_cost_terms_t2_init})", number_format="0.0")
stack_cost_terms_t2_init = ",".join(
    f'IF(ISNUMBER({COST_COL_T2[n]}{init2}),{COST_COL_T2[n]}{init2},9E15)'
    for n in STACKABLES_T2)
formula_cell(sim2, init2, col_idx(CHEAPEST_STACK_COL_T2),
             f"=MIN({stack_cost_terms_t2_init})", number_format="0.0")
formula_cell(sim2, init2, col_idx(SAVE_MODE_COL_T2), "=FALSE")

# VPCs at tick 0 = 0 (nothing to evaluate yet)
for name in STACKABLES_T2:
    formula_cell(sim2, init2, col_idx(VPC_COL_T2[name]), 0, number_format="0.000000")
formula_cell(sim2, init2, col_idx(NEXT_OS_VPC_COL_T2), 0, number_format="0.000000")
formula_cell(sim2, init2, col_idx(ACTION_COL_T2), "")


# Per-tick rows for T2
for tick in range(1, SIM2_ROWS + 1):
    r = SIM2_START + tick
    pr = r - 1

    # A: tick, B: time, C: mass start (= prev mass end), D: cohesion (= prev + prev coh delta)
    formula_cell(sim2, r, 1, tick, number_format="0")
    formula_cell(sim2, r, 2, f"=A{r}*10/60", number_format="0.00")
    formula_cell(sim2, r, 3, f"=X{pr}", number_format="0.0")
    formula_cell(sim2, r, 4, f"=D{pr}+W{pr}", number_format="0.00")

    # Levels E-M: stackables increment, one-shots MAX-cap to 1
    for n in NAMES_T2:
        col = col_idx(LVL_COL_T2[n])
        short = SHORT_T2[n]
        prev_lvl = f"{LVL_COL_T2[n]}{pr}"
        if n in STACKABLES_T2:
            formula_cell(sim2, r, col,
                         f'={prev_lvl}+IF(U{pr}="{short}",1,0)', number_format="0")
        else:
            formula_cell(sim2, r, col,
                         f'=MAX({prev_lvl},IF(U{pr}="{short}",1,0))', number_format="0")
    # N: T3 transitioned flag
    formula_cell(sim2, r, col_idx("N"),
                 f'=MAX(N{pr},IF(U{pr}="transition",1,0))', number_format="0")

    # O: MPC, P: MPS, Q: APS
    formula_cell(sim2, r, col_idx("O"),
                 aggregate_formula_T2('mpc', r, base_value=P('base_click_value')),
                 number_format="0.000", fill=DERIVED_FILL)
    formula_cell(sim2, r, col_idx("P"),
                 aggregate_formula_T2('mps', r),
                 number_format="0.000", fill=DERIVED_FILL)
    formula_cell(sim2, r, col_idx("Q"),
                 aggregate_formula_T2('aps', r),
                 number_format="0.000", fill=DERIVED_FILL)

    # R: click income, S: auto income, T: total
    # After T3 transition, sim halts income accrual
    formula_cell(sim2, r, col_idx("R"),
                 f"=IF(N{r}=1,0,O{r}*{P('active_clicks_per_min')}/60*{P('active_play_fraction')}*10)",
                 number_format="0.00")
    formula_cell(sim2, r, col_idx("S"),
                 f"=IF(N{r}=1,0,Q{r}*O{r}*10)", number_format="0.00")
    formula_cell(sim2, r, col_idx("T"),
                 f"=IF(N{r}=1,0,R{r}+S{r}+P{r}*10)", number_format="0.00")

    # ---- Helper columns ----
    for n in NAMES_T2:
        formula_cell(sim2, r, col_idx(COST_COL_T2[n]),
                     cost_formula_T2(n, r), number_format="0.0")

    # Trans? / Could-trans?
    completionist_check_t2_r = "*".join(f"({LVL_COL_T2[n]}{r}>={U(n,'E')})" for n in COMPLETIONIST_T2)
    formula_cell(sim2, r, col_idx(TRANS_COL_T2),
                 f'=IF(AND(D{r}>={t2_to_t3_cohesion},{completionist_check_t2_r}=1,N{r}=0),1,"")',
                 number_format="0")
    formula_cell(sim2, r, col_idx(COULD_TRANS_COL_T2),
                 f'=IF(AND(D{r}>={t2_to_t3_cohesion},N{r}=0),1,"")',
                 number_format="0")

    # Next OS cost / cheapest stackable
    oneshot_cost_terms_r = ",".join(
        f'IF(ISNUMBER({COST_COL_T2[n]}{r}),{COST_COL_T2[n]}{r},9E15)' for n in ONESHOTS_T2)
    formula_cell(sim2, r, col_idx(NEXT_OS_COL_T2),
                 f"=MIN({oneshot_cost_terms_r})", number_format="0.0")
    stack_cost_terms_r = ",".join(
        f'IF(ISNUMBER({COST_COL_T2[n]}{r}),{COST_COL_T2[n]}{r},9E15)' for n in STACKABLES_T2)
    formula_cell(sim2, r, col_idx(CHEAPEST_STACK_COL_T2),
                 f"=MIN({stack_cost_terms_r})", number_format="0.0")

    # Mass-after-income ref (for affordability checks) = C + T (start + total income)
    mass_avail = f"(C{r}+T{r})"

    # Post-cohesion focus: once cohesion 2.5 is met but completionists not all maxed,
    # gate off non-completionist stackables (SK, LB, ML, RLO).
    post_coh_focus = f'AND({COULD_TRANS_COL_T2}{r}=1,{TRANS_COL_T2}{r}<>1)'
    sk_aff  = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL_T2["Stellar Kinematics"]}{r}),{COST_COL_T2["Stellar Kinematics"]}{r}<={mass_avail}),{VPC_COL_T2["Stellar Kinematics"]}{r},0)'
    lb_aff  = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL_T2["Local Bubble"]}{r}),{COST_COL_T2["Local Bubble"]}{r}<={mass_avail}),{VPC_COL_T2["Local Bubble"]}{r},0)'
    ml_aff  = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL_T2["Microlensing"]}{r}),{COST_COL_T2["Microlensing"]}{r}<={mass_avail}),{VPC_COL_T2["Microlensing"]}{r},0)'
    rlo_aff = f'IF(AND(NOT({post_coh_focus}),ISNUMBER({COST_COL_T2["Roche Lobe Overflow"]}{r}),{COST_COL_T2["Roche Lobe Overflow"]}{r}<={mass_avail}),{VPC_COL_T2["Roche Lobe Overflow"]}{r},0)'
    bd_aff  = f'IF(AND(ISNUMBER({COST_COL_T2["Brown Dwarf"]}{r}),{COST_COL_T2["Brown Dwarf"]}{r}<={mass_avail}),{VPC_COL_T2["Brown Dwarf"]}{r},0)'
    max_aff_vpc_t2 = f"MAX({sk_aff},{lb_aff},{ml_aff},{rlo_aff},{bd_aff})"

    # Save mode = next-OS-VPC > threshold × max affordable VPC
    formula_cell(sim2, r, col_idx(SAVE_MODE_COL_T2),
                 f"=IF({NEXT_OS_VPC_COL_T2}{r}>{P('save_vpc_threshold')}*({max_aff_vpc_t2}),TRUE,FALSE)")

    # ---- VPC formulas for T2 stackables ----
    # Standard rule: marginal income gain per level / next purchase cost.
    # For SYNERGY PROVIDERS (RLO, BD) we additionally include the value they add to their target.
    #
    # SK (passive): per-level + M/sec × global_T2_mps × incoming_synergy × carry_all_mps
    # LB (passive): same form (incoming synergy from RLO already captured)
    # ML (click):  per-level + M/click × global_T2_mpc × incoming_synergy × carry_all_mpc × cpm × frac / 60
    # RLO (APS):   per-level + AC/sec × global_T2_aps × incoming_synergy × carry_all_aps × MPC
    #              PLUS synergy contribution to LB:
    #                LB current contribution × (1.05 - 1) ≈ LB_current × 0.05 (per RLO level marginal)
    # BD (passive): per-level + M/sec × global_T2_mps × incoming_synergy × carry_all_mps
    #              PLUS synergy contribution to RLO:
    #                RLO current contribution × (1.10 - 1) ≈ RLO_current × 0.10

    # Helpers for current per-upgrade contribution-to-income (M/sec rate)
    def own_mps_contrib_per_sec(name, r):
        # The provider's own M/sec contribution AT current state (used for synergy-gain estimation)
        base = self_contrib_formula_T2(name, 'mps', r)
        syn = synergy_mult_formula_T2(name, r)
        return f"({base})*({syn})*{global_mult_formula_T2('mps', r)}*{T2C['all_mps']}"

    def own_mpc_contrib_per_sec(name, r):
        # Provider's own M/click as a M/sec rate = M/click × cpm/60 × frac (active) + APS × M/click (auto)
        base = self_contrib_formula_T2(name, 'mpc', r)
        syn = synergy_mult_formula_T2(name, r)
        mpc_contrib = f"({base})*({syn})*{global_mult_formula_T2('mpc', r)}*{T2C['all_mpc']}"
        return f"({mpc_contrib})*({P('active_clicks_per_min')}/60*{P('active_play_fraction')}+Q{r})"

    def own_aps_contrib_per_sec(name, r):
        # APS contribution × current MPC = M/sec rate from this APS source
        base = self_contrib_formula_T2(name, 'aps', r)
        syn = synergy_mult_formula_T2(name, r)
        aps_contrib = f"({base})*({syn})*{global_mult_formula_T2('aps', r)}*{T2C['all_aps']}"
        return f"({aps_contrib})*O{r}"

    # SK VPC
    sk_delta = (f"{U('Stellar Kinematics','H')}*({global_mult_formula_T2('mps', r)})"
                f"*({synergy_mult_formula_T2('Stellar Kinematics', r)})*{T2C['all_mps']}")
    formula_cell(sim2, r, col_idx(VPC_COL_T2["Stellar Kinematics"]),
                 f'=IFERROR({sk_delta}/{COST_COL_T2["Stellar Kinematics"]}{r},0)',
                 number_format="0.000000")

    # LB VPC (target of synergy from RLO; synergy is already included via synergy_mult_formula_T2)
    lb_delta = (f"{U('Local Bubble','H')}*({global_mult_formula_T2('mps', r)})"
                f"*({synergy_mult_formula_T2('Local Bubble', r)})*{T2C['all_mps']}")
    formula_cell(sim2, r, col_idx(VPC_COL_T2["Local Bubble"]),
                 f'=IFERROR({lb_delta}/{COST_COL_T2["Local Bubble"]}{r},0)',
                 number_format="0.000000")

    # ML VPC (click-based; target of synergy from BiP)
    ml_delta = (f"{U('Microlensing','K')}*({global_mult_formula_T2('mpc', r)})"
                f"*({synergy_mult_formula_T2('Microlensing', r)})*{T2C['all_mpc']}"
                f"*({P('active_clicks_per_min')}/60*{P('active_play_fraction')}+Q{r})")
    formula_cell(sim2, r, col_idx(VPC_COL_T2["Microlensing"]),
                 f'=IFERROR({ml_delta}/{COST_COL_T2["Microlensing"]}{r},0)',
                 number_format="0.000000")

    # RLO VPC: own AC/sec contribution × current MPC + synergy gift to LB (5% of LB's current)
    rlo_self_delta = (f"{U('Roche Lobe Overflow','N')}*({global_mult_formula_T2('aps', r)})"
                      f"*({synergy_mult_formula_T2('Roche Lobe Overflow', r)})*{T2C['all_aps']}*O{r}")
    # Synergy gift to LB: current LB contribution × (1.05 - 1)
    lb_current = own_mps_contrib_per_sec("Local Bubble", r)
    rlo_synergy_gift = f"({lb_current})*0.05"
    formula_cell(sim2, r, col_idx(VPC_COL_T2["Roche Lobe Overflow"]),
                 f'=IFERROR(({rlo_self_delta}+{rlo_synergy_gift})/{COST_COL_T2["Roche Lobe Overflow"]}{r},0)',
                 number_format="0.000000")

    # BD VPC: own M/sec contribution + synergy gift to RLO (10% of RLO's current contribution)
    bd_self_delta = (f"{U('Brown Dwarf','H')}*({global_mult_formula_T2('mps', r)})"
                     f"*({synergy_mult_formula_T2('Brown Dwarf', r)})*{T2C['all_mps']}")
    rlo_current = own_aps_contrib_per_sec("Roche Lobe Overflow", r)
    bd_synergy_gift = f"({rlo_current})*0.10"
    formula_cell(sim2, r, col_idx(VPC_COL_T2["Brown Dwarf"]),
                 f'=IFERROR(({bd_self_delta}+{bd_synergy_gift})/{COST_COL_T2["Brown Dwarf"]}{r},0)',
                 number_format="0.000000")

    # ---- One-shot save VPCs (for save-mode trigger) ----
    # BiP: synergy ×1.5 on ML's self-contrib → delta = 0.5 × ML's contribution-to-M/sec
    ml_self_mpc_r = self_contrib_formula_T2("Microlensing", "mpc", r)
    ml_synergy_r = synergy_mult_formula_T2("Microlensing", r)
    ml_per_sec_contrib = (f"({ml_self_mpc_r})*({ml_synergy_r})"
                          f"*({global_mult_formula_T2('mpc', r)})*{T2C['all_mpc']}"
                          f"*({P('active_clicks_per_min')}/60*{P('active_play_fraction')}+Q{r})")
    bip_vpc = f'IF({LVL_COL_T2["Binary Partner"]}{r}>0,0,IFERROR(0.5*({ml_per_sec_contrib})/{U("Binary Partner","C")},0))'

    # PV: ×1.30 to all M/sec → delta = 0.30 × current global MPS rate
    pv_vpc = f'IF({LVL_COL_T2["Peculiar Velocity"]}{r}>0,0,IFERROR(0.30*P{r}/{U("Peculiar Velocity","C")},0))'

    # OC: pure cohesion gate. No income delta. Score is income-foregone-while-saving = 0
    # but we want the simulator to BUY it once cohesion-required. Use a tiny non-zero score
    # tied to remaining cohesion need so it ranks below productive options when cohesion already met.
    oc_vpc = f'IF({LVL_COL_T2["Open Cluster"]}{r}>0,0,IFERROR(0.0001/{U("Open Cluster","C")},0))'

    # MG: +6.0 base M/sec (passive) AND ×1.20 to all M/click
    # delta = 6.0 × global_mps × all_mps + 0.20 × current_click_M_per_sec
    mg_passive = f"6.0*({global_mult_formula_T2('mps', r)})*{T2C['all_mps']}"
    # Current click M/sec rate = MPC × cpm / 60 × active_frac + APS × MPC
    mg_click = f"0.20*O{r}*({P('active_clicks_per_min')}/60*{P('active_play_fraction')}+Q{r})"
    mg_vpc = f'IF({LVL_COL_T2["Moving Group"]}{r}>0,0,IFERROR(({mg_passive}+{mg_click})/{U("Moving Group","C")},0))'

    # Stackable save VPCs (when unaffordable)
    sk_save = f'IF(AND(ISNUMBER({COST_COL_T2["Stellar Kinematics"]}{r}),{COST_COL_T2["Stellar Kinematics"]}{r}>{mass_avail}),{VPC_COL_T2["Stellar Kinematics"]}{r},0)'
    lb_save = f'IF(AND(ISNUMBER({COST_COL_T2["Local Bubble"]}{r}),{COST_COL_T2["Local Bubble"]}{r}>{mass_avail}),{VPC_COL_T2["Local Bubble"]}{r},0)'
    ml_save = f'IF(AND(ISNUMBER({COST_COL_T2["Microlensing"]}{r}),{COST_COL_T2["Microlensing"]}{r}>{mass_avail}),{VPC_COL_T2["Microlensing"]}{r},0)'
    rlo_save = f'IF(AND(ISNUMBER({COST_COL_T2["Roche Lobe Overflow"]}{r}),{COST_COL_T2["Roche Lobe Overflow"]}{r}>{mass_avail}),{VPC_COL_T2["Roche Lobe Overflow"]}{r},0)'
    bd_save = f'IF(AND(ISNUMBER({COST_COL_T2["Brown Dwarf"]}{r}),{COST_COL_T2["Brown Dwarf"]}{r}>{mass_avail}),{VPC_COL_T2["Brown Dwarf"]}{r},0)'

    formula_cell(sim2, r, col_idx(NEXT_OS_VPC_COL_T2),
                 f"=MAX({bip_vpc},{pv_vpc},{oc_vpc},{mg_vpc},"
                 f"{sk_save},{lb_save},{ml_save},{rlo_save},{bd_save})",
                 number_format="0.000000")

    # ---- Action ----
    # Priority:
    #   1. transition if available
    #   2. buy cheapest affordable one-shot
    #   3. if save mode, do nothing
    #   4. else buy stackable with max VPC
    affordable_BiP = f'AND({LVL_COL_T2["Binary Partner"]}{r}=0,ISNUMBER({COST_COL_T2["Binary Partner"]}{r}),{COST_COL_T2["Binary Partner"]}{r}<={mass_avail})'
    affordable_PV = f'AND({LVL_COL_T2["Peculiar Velocity"]}{r}=0,ISNUMBER({COST_COL_T2["Peculiar Velocity"]}{r}),{COST_COL_T2["Peculiar Velocity"]}{r}<={mass_avail})'
    affordable_OC = f'AND({LVL_COL_T2["Open Cluster"]}{r}=0,ISNUMBER({COST_COL_T2["Open Cluster"]}{r}),{COST_COL_T2["Open Cluster"]}{r}<={mass_avail})'
    affordable_MG = f'AND({LVL_COL_T2["Moving Group"]}{r}=0,ISNUMBER({COST_COL_T2["Moving Group"]}{r}),{COST_COL_T2["Moving Group"]}{r}<={mass_avail})'
    cheapest_affordable_os_t2 = (
        f"MIN("
        f"IF({affordable_BiP},{COST_COL_T2['Binary Partner']}{r},9E15),"
        f"IF({affordable_PV},{COST_COL_T2['Peculiar Velocity']}{r},9E15),"
        f"IF({affordable_OC},{COST_COL_T2['Open Cluster']}{r},9E15),"
        f"IF({affordable_MG},{COST_COL_T2['Moving Group']}{r},9E15)"
        f")")
    pick_BiP = f'AND({affordable_BiP},{COST_COL_T2["Binary Partner"]}{r}={cheapest_affordable_os_t2})'
    pick_PV = f'AND({affordable_PV},{COST_COL_T2["Peculiar Velocity"]}{r}={cheapest_affordable_os_t2})'
    pick_OC = f'AND({affordable_OC},{COST_COL_T2["Open Cluster"]}{r}={cheapest_affordable_os_t2})'

    # Stackable selection
    pick_SK = f'AND({sk_aff}>0,{sk_aff}={max_aff_vpc_t2})'
    pick_LB = f'AND({lb_aff}>0,{lb_aff}={max_aff_vpc_t2})'
    pick_ML = f'AND({ml_aff}>0,{ml_aff}={max_aff_vpc_t2})'
    pick_RLO = f'AND({rlo_aff}>0,{rlo_aff}={max_aff_vpc_t2})'

    formula_cell(sim2, r, col_idx(ACTION_COL_T2),
                 f'=IF({TRANS_COL_T2}{r}=1,"transition",'
                 f'IF({cheapest_affordable_os_t2}<9E15,'
                   f'IF({pick_BiP},"BiP",IF({pick_PV},"PV",IF({pick_OC},"OC","MG"))),'
                 f'IF({SAVE_MODE_COL_T2}{r}=TRUE,"",'
                 f'IF({max_aff_vpc_t2}<=0,"",'
                 f'IF({pick_SK},"SK",'
                 f'IF({pick_LB},"LB",'
                 f'IF({pick_ML},"ML",'
                 f'IF({pick_RLO},"RLO","BD"))))))))')

    # U: Buy?
    formula_cell(sim2, r, col_idx("U"), f"={ACTION_COL_T2}{r}",
                 font=EVENT_FONT, alignment=CENTER)
    # V: Cost (lookup based on action code)
    cost_lookup_t2 = (
        f'=IF(U{r}="",0,IF(U{r}="transition",0,'
        f'IF(U{r}="SK",{COST_COL_T2["Stellar Kinematics"]}{r},'
        f'IF(U{r}="LB",{COST_COL_T2["Local Bubble"]}{r},'
        f'IF(U{r}="ML",{COST_COL_T2["Microlensing"]}{r},'
        f'IF(U{r}="RLO",{COST_COL_T2["Roche Lobe Overflow"]}{r},'
        f'IF(U{r}="BD",{COST_COL_T2["Brown Dwarf"]}{r},'
        f'IF(U{r}="BiP",{COST_COL_T2["Binary Partner"]}{r},'
        f'IF(U{r}="PV",{COST_COL_T2["Peculiar Velocity"]}{r},'
        f'IF(U{r}="OC",{COST_COL_T2["Open Cluster"]}{r},'
        f'IF(U{r}="MG",{COST_COL_T2["Moving Group"]}{r},0))))))))))'
    )
    formula_cell(sim2, r, col_idx("V"), cost_lookup_t2, number_format="0.0")
    # W: Cohesion delta
    formula_cell(sim2, r, col_idx("W"),
                 f'=IF(U{r}="BiP",{U("Binary Partner","F")},'
                 f'IF(U{r}="PV",{U("Peculiar Velocity","F")},'
                 f'IF(U{r}="OC",{U("Open Cluster","F")},'
                 f'IF(U{r}="MG",{U("Moving Group","F")},'
                 f'IF(U{r}="transition",-{t2_to_t3_cohesion},0)))))',
                 number_format="0.00")
    # X: mass end = (start + total income) - cost
    formula_cell(sim2, r, col_idx("X"), f"=C{r}+T{r}-V{r}", number_format="0.0")
    # Y: M/sec rate
    formula_cell(sim2, r, col_idx("Y"),
                 f"=O{r}*{P('active_clicks_per_min')}/60*{P('active_play_fraction')}+"
                 f"Q{r}*O{r}+P{r}",
                 number_format="0.00", fill=DERIVED_FILL)
    # Z: note
    formula_cell(sim2, r, col_idx("Z"),
                 f'=IF(U{r}="transition","→ Tier 3 (Galactic Arm)",'
                 f'IF(U{r}="MG","Moving Group — pocket drifts together",'
                 f'IF(U{r}="OC","Open Cluster acquired",'
                 f'IF(U{r}="PV","Peculiar Velocity — passive multiplier active",'
                 f'IF(U{r}="BiP","Binary Partner — pair bond formed",'
                 f'IF(AND({COULD_TRANS_COL_T2}{r}=1,N{r}=0,{TRANS_COL_T2}{r}<>1),'
                 f'"Cohesion ready (waiting on completionists)","")))))) ',
                 font=EVENT_FONT, alignment=LEFT_NOWRAP)

    # Per-upgrade contribution displays AA-AJ
    for i, name in enumerate(NAMES_T2):
        # Determine the dominant stat for this upgrade's display
        # SK, LB, BD, MG: M/sec passive
        # ML: M/click → display in M/min via cpm
        # RLO: AC/sec × MPC → display in M/min
        # BiP, PV, OC: zero direct contribution
        col_map_idx = 27 + i  # AA = 27, AB = 28, ..., AJ = 36
        if name in ["Stellar Kinematics", "Local Bubble", "Brown Dwarf"]:
            base = self_contrib_formula_T2(name, "mps", r)
            syn = synergy_mult_formula_T2(name, r)
            formula_cell(sim2, r, col_map_idx,
                         f"=({base})*({syn})*{global_mult_formula_T2('mps', r)}*{T2C['all_mps']}*60",
                         number_format="0.00")
        elif name == "Microlensing":
            base = self_contrib_formula_T2(name, "mpc", r)
            syn = synergy_mult_formula_T2(name, r)
            formula_cell(sim2, r, col_map_idx,
                         f"=({base})*({syn})*{global_mult_formula_T2('mpc', r)}*{T2C['all_mpc']}*"
                         f"({P('active_clicks_per_min')}*{P('active_play_fraction')}+Q{r}*60)",
                         number_format="0.00")
        elif name == "Roche Lobe Overflow":
            # AC/sec × MPC × 60
            base = self_contrib_formula_T2(name, "aps", r)
            syn = synergy_mult_formula_T2(name, r)
            formula_cell(sim2, r, col_map_idx,
                         f"=({base})*({syn})*{global_mult_formula_T2('aps', r)}*{T2C['all_aps']}*O{r}*60",
                         number_format="0.00")
        elif name == "Moving Group":
            # base 6.0 M/sec — handled by self_contrib_formula since it's in the MPS base column
            base = self_contrib_formula_T2(name, "mps", r)
            syn = synergy_mult_formula_T2(name, r)
            formula_cell(sim2, r, col_map_idx,
                         f"=({base})*({syn})*{global_mult_formula_T2('mps', r)}*{T2C['all_mps']}*60",
                         number_format="0.00")
        else:
            # BiP, PV, OC — direct contribution is 0
            formula_cell(sim2, r, col_map_idx, 0, number_format="0.00")

    # AJ (col 36): Carry M/min display = carry MPS × 60 (passive only; click contribution flows
    # through MPC rather than appearing as a direct M/sec component)
    formula_cell(sim2, r, col_idx("AJ"),
                 f"={T2C['mps']}*60", number_format="0.00")


# Headline cells for T2
sim2.cell(row=T2_HEADLINE_ROW, column=1,
          value="Earliest possible T3 transition (cohesion ready):").font = NOTE_FONT
sim2.cell(row=T2_HEADLINE_ROW, column=1).alignment = LEFT_NOWRAP
sim2.merge_cells(start_row=T2_HEADLINE_ROW, start_column=1,
                 end_row=T2_HEADLINE_ROW, end_column=4)
hl_t2_a = sim2.cell(row=T2_HEADLINE_ROW, column=5,
                    value=f'=IFERROR(INDEX(B:B,MATCH(1,{COULD_TRANS_COL_T2}:{COULD_TRANS_COL_T2},0))&" min","—")')
hl_t2_a.font = LINK_FONT
hl_t2_a.alignment = LEFT_NOWRAP
sim2.cell(row=T2_HEADLINE_ROW, column=8,
          value="Actual T3 transition (completion path):").font = NOTE_FONT
sim2.cell(row=T2_HEADLINE_ROW, column=8).alignment = LEFT_NOWRAP
sim2.merge_cells(start_row=T2_HEADLINE_ROW, start_column=8,
                 end_row=T2_HEADLINE_ROW, end_column=13)
hl_t2_b = sim2.cell(row=T2_HEADLINE_ROW, column=14,
                    value=f'=IFERROR(INDEX(B:B,MATCH("transition",U:U,0))&" min","—")')
hl_t2_b.font = LINK_FONT
hl_t2_b.alignment = LEFT_NOWRAP

sim2.freeze_panes = "C11"


# ============================================================
# Curves — purchase cadence + click share diagnostic
# ============================================================
curves = wb.create_sheet("Curves")
set_widths(curves, {"A": 8, "B": 12, "C": 14, "D": 14, "E": 14})

ct = curves.cell(row=1, column=1, value="Curves")
ct.font = H1
ct.alignment = LEFT_NOWRAP

cs = curves.cell(row=2, column=1,
                 value="Purchase cadence and click share diagnostics for T1.")
cs.font = NOTE_FONT
cs.alignment = LEFT_NOWRAP

# Purchase cadence: per-minute purchase count
header_cell(curves, 4, 1, "Minute")
header_cell(curves, 4, 2, "Purchases")
header_cell(curves, 4, 3, "Cum purchases")
header_cell(curves, 4, 4, "Click share %")
header_cell(curves, 4, 5, "Click share target")

# 30 minutes max for T1 chart
NUM_BUCKETS = 30
for m in range(NUM_BUCKETS):
    r = 5 + m
    formula_cell(curves, r, 1, m, number_format="0", font=LINK_FONT)
    # Purchases in this minute = COUNTIFS where time_min ∈ [m, m+1) AND Buy? has text
    # In v1.1, the Buy? column moved from S to U.
    # Note: criteria "?*" matches any cell with at least one character of text.
    formula_cell(curves, r, 2,
                 f'=COUNTIFS(UpgradeSim!$B$5:$B$405,">="&{m},'
                 f'UpgradeSim!$B$5:$B$405,"<"&{m+1},'
                 f'UpgradeSim!$U$5:$U$405,"?*")',
                 number_format="0", font=LINK_FONT)
    # Cumulative
    formula_cell(curves, r, 3, f"=SUM(B$5:B{r})", number_format="0", font=LINK_FONT)
    # Click share at the closest tick within this minute
    # In v1.1: Click income/tick is in column P, Total income/tick in column S.
    formula_cell(curves, r, 4,
                 f'=IFERROR(INDEX(UpgradeSim!$P$5:$P$405,MATCH({m+1},UpgradeSim!$B$5:$B$405,1))/'
                 f'(INDEX(UpgradeSim!$S$5:$S$405,MATCH({m+1},UpgradeSim!$B$5:$B$405,1))),0)',
                 number_format="0%", font=LINK_FONT)
    # Click share target (T1 only — uses T1 click share parameter as constant target)
    formula_cell(curves, r, 5, f"={P('click_share_T1')}",
                 number_format="0%", font=LINK_FONT)

# Charts
ch1 = BarChart()
ch1.type = "col"
ch1.style = 10
ch1.title = "T1 Purchases per Minute"
ch1.y_axis.title = "Purchases"
ch1.x_axis.title = "Minute"
ch1.add_data(Reference(curves, min_col=2, min_row=4, max_row=4 + NUM_BUCKETS),
             titles_from_data=True)
ch1.set_categories(Reference(curves, min_col=1, min_row=5, max_row=4 + NUM_BUCKETS))
ch1.height = 9
ch1.width = 16
curves.add_chart(ch1, "G4")

ch2 = LineChart()
ch2.title = "T1 Click Share — actual vs target"
ch2.y_axis.title = "Click share"
ch2.x_axis.title = "Minute"
ch2.style = 12
ch2.add_data(Reference(curves, min_col=4, min_row=4, max_row=4 + NUM_BUCKETS),
             titles_from_data=True)
ch2.add_data(Reference(curves, min_col=5, min_row=4, max_row=4 + NUM_BUCKETS),
             titles_from_data=True)
ch2.set_categories(Reference(curves, min_col=1, min_row=5, max_row=4 + NUM_BUCKETS))
ch2.height = 9
ch2.width = 16
curves.add_chart(ch2, "G24")


# ============================================================
# CurvesT2 — T2 purchase cadence + click share diagnostic
# ============================================================
curves2 = wb.create_sheet("CurvesT2")
set_widths(curves2, {"A": 8, "B": 12, "C": 14, "D": 14, "E": 14, "F": 14})

ct2 = curves2.cell(row=1, column=1, value="Curves — Tier 2")
ct2.font = H1
ct2.alignment = LEFT_NOWRAP

cs2 = curves2.cell(row=2, column=1,
                   value="Purchase cadence and income split for T2. Tracks T2 sim minutes (independent of T1 minutes).")
cs2.font = NOTE_FONT
cs2.alignment = LEFT_NOWRAP

header_cell(curves2, 4, 1, "Minute (T2)")
header_cell(curves2, 4, 2, "Purchases")
header_cell(curves2, 4, 3, "Cum purchases")
header_cell(curves2, 4, 4, "Click share %")
header_cell(curves2, 4, 5, "Active income %")
header_cell(curves2, 4, 6, "Auto income %")

# T2 sim sheet rows: data starts at row 11 (init) and runs +400 ticks (~67 min)
T2_SIM_ROW_START = SIM2_START
T2_SIM_ROW_END = SIM2_START + SIM2_ROWS  # 11 + 400 = 411

NUM_BUCKETS_T2 = 40  # cover ~40 minutes of T2 progression
for m in range(NUM_BUCKETS_T2):
    r = 5 + m
    formula_cell(curves2, r, 1, m, number_format="0", font=LINK_FONT)
    # Purchases this minute (T2 Buy column = U on sim2 sheet)
    formula_cell(curves2, r, 2,
                 f'=COUNTIFS(UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},">="&{m},'
                 f'UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},"<"&{m+1},'
                 f'UpgradeSimT2!$U${T2_SIM_ROW_START}:$U${T2_SIM_ROW_END},"?*")',
                 number_format="0", font=LINK_FONT)
    formula_cell(curves2, r, 3, f"=SUM(B$5:B{r})", number_format="0", font=LINK_FONT)
    # Click share = R / T (click income / total income) at first tick of this minute
    formula_cell(curves2, r, 4,
                 f'=IFERROR(INDEX(UpgradeSimT2!$R${T2_SIM_ROW_START}:$R${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1))/'
                 f'INDEX(UpgradeSimT2!$T${T2_SIM_ROW_START}:$T${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1)),0)',
                 number_format="0%", font=LINK_FONT)
    # Active income % = (click + autoclick) / total
    formula_cell(curves2, r, 5,
                 f'=IFERROR((INDEX(UpgradeSimT2!$R${T2_SIM_ROW_START}:$R${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1))+'
                 f'INDEX(UpgradeSimT2!$S${T2_SIM_ROW_START}:$S${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1)))/'
                 f'INDEX(UpgradeSimT2!$T${T2_SIM_ROW_START}:$T${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1)),0)',
                 number_format="0%", font=LINK_FONT)
    # Auto income % = autoclick / total
    formula_cell(curves2, r, 6,
                 f'=IFERROR(INDEX(UpgradeSimT2!$S${T2_SIM_ROW_START}:$S${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1))/'
                 f'INDEX(UpgradeSimT2!$T${T2_SIM_ROW_START}:$T${T2_SIM_ROW_END},'
                 f'MATCH({m+1},UpgradeSimT2!$B${T2_SIM_ROW_START}:$B${T2_SIM_ROW_END},1)),0)',
                 number_format="0%", font=LINK_FONT)

# T2 charts
ch1_t2 = BarChart()
ch1_t2.type = "col"
ch1_t2.style = 10
ch1_t2.title = "T2 Purchases per Minute"
ch1_t2.y_axis.title = "Purchases"
ch1_t2.x_axis.title = "Minute (T2)"
ch1_t2.add_data(Reference(curves2, min_col=2, min_row=4,
                          max_row=4 + NUM_BUCKETS_T2), titles_from_data=True)
ch1_t2.set_categories(Reference(curves2, min_col=1, min_row=5,
                                max_row=4 + NUM_BUCKETS_T2))
ch1_t2.height = 9
ch1_t2.width = 16
curves2.add_chart(ch1_t2, "H4")

ch2_t2 = LineChart()
ch2_t2.title = "T2 Income Split (active vs auto)"
ch2_t2.y_axis.title = "Share"
ch2_t2.x_axis.title = "Minute (T2)"
ch2_t2.style = 12
ch2_t2.add_data(Reference(curves2, min_col=4, min_row=4,
                          max_row=4 + NUM_BUCKETS_T2), titles_from_data=True)
ch2_t2.add_data(Reference(curves2, min_col=5, min_row=4,
                          max_row=4 + NUM_BUCKETS_T2), titles_from_data=True)
ch2_t2.add_data(Reference(curves2, min_col=6, min_row=4,
                          max_row=4 + NUM_BUCKETS_T2), titles_from_data=True)
ch2_t2.set_categories(Reference(curves2, min_col=1, min_row=5,
                                max_row=4 + NUM_BUCKETS_T2))
ch2_t2.height = 9
ch2_t2.width = 16
curves2.add_chart(ch2_t2, "H24")


# Reorder
ordering = ["README", "Parameters", "Upgrades", "TimeBudget",
            "UpgradeSim", "Curves", "UpgradeSimT2", "CurvesT2"]
for i, name in enumerate(ordering):
    wb.move_sheet(name, offset=i - wb.index(wb[name]))

wb.active = wb.index(wb["UpgradeSimT2"])

import os
out_dir = os.path.dirname(os.path.abspath(__file__))
out_path = os.path.join(out_dir, "dark-filaments-simulation-v1.3.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
